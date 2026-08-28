"""مرکز شارژ داده — Bulk Data Charge Module (§6B MODULE EQUIPMENT).

Workflow:
    upload raw workbook → staging (never straight into main DB)
    → SELEN assisted mapping (user confirms) → validation + fuzzy duplicate
    detection → diff preview (New/Update/Conflict/Rejected) → manual fixes
    → Commit (clean rows only) → batch Rollback.

Raw source files are kept for audit; every step is audit-logged (§6B AUDIT).
"""
from __future__ import annotations

import io
from datetime import datetime, timezone
from difflib import SequenceMatcher

from fastapi import APIRouter, Depends, File, HTTPException, Request, UploadFile
from openpyxl import load_workbook
from pydantic import BaseModel
from sqlalchemy.orm import Session

from .. import audit, storage
from ..ai.equipment_codes import (
    EQUIPMENT_FIELDS, HEADER_ALIASES, decode_code, guess_field_from_value,
    is_date_like, suggest_mapping,
)
from ..db import get_db, utcnow
from ..events import bus
from ..models import (
    Equipment, EquipmentCategory, Factory, ImportBatch, ImportBatchRow,
    LookupItem, Part, User,
)
from ..rbac import require

router = APIRouter(prefix="/equipment/bulk-charge", tags=["bulk-charge"])

CRIT_ALIASES = {
    "کم": "low", "پایین": "low", "متوسط": "medium", "زیاد": "high",
    "بالا": "high", "بحرانی": "critical",
    "a": "critical", "b": "high", "c": "medium", "d": "low",
    "low": "low", "medium": "medium", "high": "high", "critical": "critical",
}
STATUS_ALIASES = {
    "فعال": "active", "غیرفعال": "inactive", "در دست تعمیر": "under_maintenance",
    "اسقاط": "scrapped", "از رده خارج": "scrapped",
    "active": "active", "inactive": "inactive",
}
REQUIRED = ("code", "name", "factory", "category")

# آستانه‌های تبدیل امتیاز عددی بحرانی (مجموع چهار شاخص
# Safety+Product+Cost+Repair در جدول DegresOFEquipment — حداکثر ۱۰۰)
CRITICALITY_THRESHOLDS = ((75, "critical"), (50, "high"), (25, "medium"), (0, "low"))


def criticality_from_score(score: float) -> str:
    for threshold, level in CRITICALITY_THRESHOLDS:
        if score >= threshold:
            return level
    return "low"


# تاریخ معادل برج ۱۰ سال ۱۴۰۴ (قاعده زمانی انتساب سوابق)
_J10_1404 = None


def _j10_1404():
    global _J10_1404
    if _J10_1404 is None:
        from ..jalali import jalali_to_gregorian
        g = jalali_to_gregorian(1404, 10, 1)
        _J10_1404 = datetime(g.year, g.month, g.day, tzinfo=timezone.utc)
    return _J10_1404


def _assign_history_technician(db, eq, title: str, at):
    """قانون نیروی انسانی — بخش ۵ سند بارگذاری نهایی.

    ماشین‌آلات کارخانه فوم (بسپار۱): برق ← کاووسی · روغن گیربکس/کاسه‌نمد ←
    پیرایش · جوشکاری ← شاه‌کرمی · بقیه مکانیک/عمومی ← تیم محلی (بعد از
    برج ۱۰ ۱۴۰۴: معافی‌پور، قبل از آن: بابایی).
    سایر دسته‌ها/کارخانه‌ها: تیم نت مرکزی — مجری پیش‌فرض تعیین نمی‌شود
    (بدون حدس؛ در پیش‌نمایش قابل ویرایش توسط کاربر).
    """
    cat = eq.category.name if eq.category else ""
    fac = eq.factory.name if eq.factory else ""
    if cat != "ماشین‌آلات تولید" or fac != "بسپار۱":
        return None
    names = ("a.kavousi", "m.pirayesh", "e.shahkarami", "n.babaei", "p.moafipour")
    users = {u.username: u for u in
             db.query(User).filter(User.username.in_(names)).all()}
    t = (title or "").lower()
    if any(k in t for k in ("برق", "وایرینگ", "تابلو", "سنسور", "کابل",
                            "کنتاکتور", "اینورتر", "سوکت", "سیم‌بندی", "سیم بندی")):
        return users.get("a.kavousi")
    if any(k in t for k in ("روغن گیربکس", "کاسه نمد", "کاسه‌نمد")):
        return users.get("m.pirayesh")
    if "جوش" in t:
        return users.get("e.shahkarami")
    if at is not None and at >= _j10_1404():
        return users.get("p.moafipour")
    return users.get("n.babaei")


def extra_dyn_fields(raw: dict) -> dict:
    """فیلدهای پویای اضافی از خروجی اکسس (ساعت کار، ابعاد، وزن، خط محصول…).
    مقادیر صفر/خالی به‌عنوان «ثبت‌نشده» در نظر گرفته می‌شوند."""
    dyn = {}
    if raw.get("install_date"):
        dyn["تاریخ نصب"] = raw["install_date"]
    if raw.get("automation"):
        dyn["درجه اتوماسیون"] = raw["automation"]
    if raw.get("equipment_type"):
        dyn["نوع تجهیز (مبدأ)"] = raw["equipment_type"]
    if raw.get("product_line"):
        dyn["خط محصول"] = raw["product_line"]
    dh = raw.get("daily_hours")
    if dh not in (None, "", "0"):
        dyn["ساعت کار روزانه"] = dh
    dims = []
    for k, lbl in (("length", "طول"), ("width", "عرض"), ("height", "ارتفاع")):
        v = raw.get(k)
        if v not in (None, "", "0"):
            dims.append(f"{lbl} {v}m")
    if dims:
        dyn["ابعاد"] = " × ".join(dims)
    w = raw.get("weight")
    if w not in (None, "", "0"):
        dyn["وزن (تن)"] = w
    return dyn


def resolve_criticality(value) -> tuple[str, int | None]:
    """متن/حرف → سطح استاندارد؛ عدد → سطحِ مبتنی بر امتیاز + حفظ امتیاز."""
    if value is None:
        return "medium", None
    s = str(value).strip()
    if not s:
        return "medium", None
    try:
        score = float(s.translate(str.maketrans("۰۱۲۳۴۵۶۷۸۹", "0123456789")))
        return criticality_from_score(score), int(score)
    except ValueError:
        pass
    return CRIT_ALIASES.get(s.lower(), "medium"), None


def _norm(v) -> str:
    return "" if v is None else str(v).strip()


def _is_num(v) -> bool:
    try:
        float(v)
        return True
    except (TypeError, ValueError):
        return False


def _parse_flexible_date(s: str):
    """Accept Jalali (1404/05/26), Gregorian (2025-08-17) and Access's
    compact YYYYMMDD (13940726). «0»/empty → None."""
    from datetime import datetime, timezone
    from ..jalali import jalali_to_gregorian

    s = (s or "").strip().replace("-", "/").replace(".", "/")
    if s in ("", "0", "0/0/0"):
        return None
    # فرمت فشرده اکسس: ۸ رقم بدون جداکننده
    if len(s) == 8 and s.isdigit():
        s = f"{s[0:4]}/{s[4:6]}/{s[6:8]}"
    parts = [p for p in s.split("/") if p.strip() != ""]
    if len(parts) != 3 or not all(p.isdigit() for p in parts):
        return None
    a, b, c = (int(p) for p in parts)
    try:
        if 1200 < a < 1500:                      # Jalali year first
            g = jalali_to_gregorian(a, b, c)
        elif a > 1900:                            # Gregorian year first
            g = __import__("datetime").date(a, b, c)
        elif c > 1900:                            # day/month/year Gregorian
            g = __import__("datetime").date(c, b, a)
        else:
            return None
        return datetime.combine(g, datetime.min.time(), tzinfo=timezone.utc)
    except (ValueError, OverflowError):
        return None


INTERVAL_FA_TO_CODE = {
    "روزانه": "daily", "هفتگی": "weekly", "دوهفتگی": "biweekly", "دو هفته": "biweekly",
    "سه‌هفته": "3weekly", "سه هفته": "3weekly", "ماهانه": "monthly",
    "دوماهه": "2monthly", "دو ماهه": "2monthly", "سه‌ماهه": "3monthly",
    "سه ماهه": "3monthly", "شش‌ماهه": "6monthly", "شش ماهه": "6monthly",
    "سالانه": "yearly", "دوسالانه": "2yearly", "دو ساله": "2yearly",
    "سفارشی": "custom",
}


def _interval_from(db, value: str):
    """Return (code, days) from a Persian title, code or plain day count."""
    v = (value or "").strip()
    if not v:
        return None, None
    if _is_num(v):  # plain number → custom days
        return "custom", int(float(v))
    key = v.replace("‌", " ").strip()
    if key.lower() in INTERVAL_FA_TO_CODE:
        key = INTERVAL_FA_TO_CODE[key.lower()]
    item = (db.query(LookupItem)
            .filter(LookupItem.list_code == "interval", LookupItem.code == key)
            .one_or_none())
    if item is None:
        item = (db.query(LookupItem)
                .filter(LookupItem.list_code == "interval",
                        LookupItem.title_fa == value.strip()).one_or_none())
    if item is None:
        return None, None
    return item.code, int((item.extra or {}).get("days", 30) or 30)


ACTIVITY_FA_TO_CODE = {
    "بازرسی": "inspection", "بازدید": "inspection",
    "تعویض قطعه": "part_replacement", "تعویض": "part_replacement",
    "تعویض روغن": "oil_change", "روانکاری": "lubrication", "روغن‌کاری": "lubrication",
    "تمیزکاری": "cleaning", "نظافت": "cleaning", "آچارکشی": "tightening",
    "تنظیم": "adjustment", "کنترل": "control", "سرویس": "service", "سایر": "other",
}


def _activity_from(db, value: str) -> str:
    v = (value or "").strip()
    if not v:
        return "inspection"
    code = ACTIVITY_FA_TO_CODE.get(v.replace("‌", " ").strip()) or v.lower()
    item = (db.query(LookupItem)
            .filter(LookupItem.list_code == "activity_type",
                    LookupItem.code == code).one_or_none())
    if item is None:
        item = (db.query(LookupItem)
                .filter(LookupItem.list_code == "activity_type",
                        LookupItem.title_fa == v).one_or_none())
    return item.code if item else "other"


def _sheet_by_hint(wb, hints: tuple[str, ...], fallback_index: int = -1):
    """Find a sheet by name hint only — positional guessing corrupts
    arbitrary legacy workbooks, so it is deliberately NOT used."""
    for ws in wb.worksheets:
        title = (ws.title or "").strip().lower()
        if any(h in title for h in hints):
            return ws
    return None


def _find_header_grid(grid: list[list[str]], max_scan: int = 8) -> tuple[int, list[str]]:
    """First row containing ≥2 known equipment headers."""
    from ..ai.equipment_codes import HEADER_ALIASES

    best = (0, [])
    for idx, row in enumerate(grid):
        if idx > max_scan:
            break
        cells = [_norm(c) for c in row]
        hits = sum(1 for c in cells if c and (c.lower() in HEADER_ALIASES or c in HEADER_ALIASES))
        if hits >= 2:
            return idx, cells
        if hits > len(best[1]):
            best = (idx, cells)
    return best


def _is_section_row(cells: list[str]) -> bool:
    """§6B: a row carrying only a group name/code — context, not a record."""
    non_empty = [c for c in cells if c]
    return 0 < len(non_empty) <= 2


def _similarity(a: str, b: str) -> float:
    if not a or not b:
        return 0.0
    return SequenceMatcher(None, a.lower(), b.lower()).ratio()


# ---------------------------------------------------------------------------
# Multi-sheet template (§6B)
# ---------------------------------------------------------------------------

@router.get("/template")
def charge_template(_: object = Depends(require("bulk_charge.charge"))):
    """6-sheet standard template: equipment, specs, structure, parts,
    PM plans, maintenance history."""
    from openpyxl import Workbook
    from fastapi.responses import StreamingResponse

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "تجهیزات"
    ws1.append(["کد تجهیز", "نام تجهیز", "کارخانه", "دسته", "سازنده", "مدل",
                "شماره سریال", "سال ساخت", "درجه اهمیت", "وضعیت", "سالن", "خط"])
    ws1.append(["B1PT-001", "پمپ سانتریفیوژ P-101", "بسپار۱", "تولیدی",
                "KSB", "ETA-65", "SN-1234", 2019, "زیاد", "فعال", "سالن ۱", "خط A"])

    ws2 = wb.create_sheet("مشخصات فنی")
    ws2.append(["کد تجهیز", "نام مشخصه", "مقدار", "واحد"])
    ws2.append(["B1PT-001", "توان", "7.5", "kW"])
    ws2.append(["B1PT-001", "دبی", "120", "m3/h"])

    ws3 = wb.create_sheet("ساختار")
    ws3.append(["کد تجهیز", "سطح", "کد والد", "نام", "سازنده", "مدل", "سریال"])
    ws3.append(["B1PT-001", "زیرسیستم", "B1PT-001", "سیستم محرک", "", "", ""])

    ws4 = wb.create_sheet("قطعات")
    ws4.append(["کد تجهیز", "کد قطعه", "نام قطعه", "موجودی فعلی", "حد موجودی",
                "مقدار سفارش", "تأمین‌کننده", "درجه اهمیت"])
    ws4.append(["B1PT-001", "PN-501", "مکانیکال سیل", 2, 1, 2, "فلان تأمین", "زیاد"])

    ws5 = wb.create_sheet("برنامه نگهداری")
    ws5.append(["کد تجهیز", "عنوان فعالیت", "نوع فعالیت", "تناوب", "مجری",
                "مدت (دقیقه)", "آخرین اجرا"])
    ws5.append(["B1PT-001", "روانکاری یاتاقان‌ها", "روانکاری", "ماهانه",
                "تیم نت", 45, "1404/04/15"])

    ws6 = wb.create_sheet("سوابق تعمیرات")
    ws6.append(["کد تجهیز", "تاریخ", "نوع کار", "عنوان", "اقدام تعمیراتی",
                "تکنسین", "مدت (دقیقه)", "هزینه (ریال)"])
    ws6.append(["B1PT-001", "1403/11/20", "تعمیر", "نشتی مکانیکال سیل",
                "تعویض مکانیکال سیل", "", 120, 8500000])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition":
                 "attachment; filename=bulk-charge-template-6sheet.xlsx"})


# ---------------------------------------------------------------------------
# Upload → staging
# ---------------------------------------------------------------------------

@router.post("/upload", status_code=201)
async def upload_charge(request: Request,
                        file: UploadFile = File(...),
                        user: User = Depends(require("bulk_charge.charge")),
                        db: Session = Depends(get_db)):
    import csv as _csvmod

    fname = (file.filename or "").lower()
    content = await file.read()
    if len(content) > 20 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="حجم فایل بیش از ۲۰ مگابایت است")

    wb = None
    if fname.endswith(".csv"):
        # Legacy CSV: UTF-8 (BOM optional) with Windows-1256 fallback and
        # auto-detected delimiter (, ; tab) — Persian exports vary wildly.
        try:
            text = content.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = content.decode("cp1256", errors="replace")
        first_line = text.splitlines()[0] if text else ""
        delim = max([",", ";", "\t"], key=lambda d: first_line.count(d))
        grid = [row for row in _csvmod.reader(io.StringIO(text), delimiter=delim)]
        grid = [[_norm(c) for c in row] for row in grid if any(_norm(c) for c in row)]
        if not grid:
            raise HTTPException(status_code=400, detail="فایل CSV خالی است")
    elif fname.endswith((".xlsx", ".xls")):
        try:
            wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        except Exception:
            raise HTTPException(status_code=400, detail="فایل Excel قابل خواندن نیست")
    else:
        raise HTTPException(status_code=400, detail="فقط فایل Excel یا CSV مجاز است")

    if wb is not None:
        # Real BASPAR workbook: several sheets, ONE equipment per sheet
        # (complete info in each).  Also tolerates a single tabular sheet.
        headers, records = _records_from_workbook(wb)
    else:
        headers, records = _records_from_grid(grid)

    header_idx = 0
    mapping = suggest_mapping(headers)

    # keep the raw file for traceability (§6B audit)
    raw_meta = await storage.save_upload(
        UploadFile(filename=file.filename, file=io.BytesIO(content)),
        entity_type="bulk_charge", entity_id="raw",
    )

    batch = ImportBatch(
        entity_type="bulk_charge", filename=file.filename or "charge",
        status="pending", mapping={"columns": mapping, "header_row": header_idx},
        raw_file_path=raw_meta["path"], created_by=user.id,
    )
    db.add(batch)
    db.flush()

    _stage_equipment_records(db, batch, records)
    if wb is not None:
        _stage_extra_sheets(db, batch, wb)

    db.flush()
    all_rows = db.query(ImportBatchRow).filter(ImportBatchRow.batch_id == batch.id).all()
    batch.total_rows = len([r for r in all_rows if r.raw.get("_sheet") in (None, "equipment")])
    audit.record(db, user_id=user.id, action="bulkcharge.uploaded",
                 entity_type="import_batch", entity_id=batch.id,
                 new={"filename": batch.filename, "rows": batch.total_rows},
                 request=request)
    db.commit()

    return {
        "batch_id": batch.id,
        "total_rows": batch.total_rows,
        "mapping": mapping,
        "header_row": header_idx,
        "message": "داده در Staging قرار گرفت؛ Mapping پیشنهادی SELEN را تأیید یا اصلاح کنید",
    }


_EXTRA_MARKERS = {"نام مشخصه", "مقدار", "واحد", "عنوان فعالیت", "تناوب",
                  "نوع کار", "part number", "نام قطعه", "سطح", "کد والد",
                  "هزینه (ریال)", "آخرین اجرا"}


def _is_extra_sheet(headers) -> bool:
    low = [(h or "").strip().lower() for h in headers if h]
    return any(m in low for m in _EXTRA_MARKERS)


def _colmap(headers) -> dict:
    out = {}
    for i, hh in enumerate(headers):
        f = HEADER_ALIASES.get(hh) or HEADER_ALIASES.get((hh or "").strip().lower())
        if f:
            out[i] = f
    return out


def _records_from_grid(grid):
    """Extract equipment records from ONE tabular grid."""
    hi, headers = _find_header_grid(grid)
    colmap = _colmap(headers)
    recs = []
    section_ctx = None
    for row in grid[hi + 1:]:
        cells = [_norm(c) for c in row]
        if not any(cells):
            continue
        if _is_section_row(cells):
            section_ctx = next((c for c in cells if c), None)
            continue
        rec = {"_section": section_ctx}
        for i, val in enumerate(cells):
            if not val:
                continue
            f = colmap.get(i)
            if f:
                rec[f] = val
            else:
                g = guess_field_from_value(val)
                if g and g not in rec:
                    rec[g] = val
        if rec.get("code") or rec.get("name"):
            recs.append(rec)
    return headers, recs


def _records_from_workbook(wb):
    """Real BASPAR format: several sheets, ONE equipment per sheet (complete
    info each).  Also tolerates a single tabular sheet with many rows."""
    all_recs = []
    headers = []
    for ws in wb.worksheets:
        grid = [[_norm(c) for c in row] for row in ws.iter_rows(values_only=True)]
        grid = [r for r in grid if any(r)]
        if not grid:
            continue
        hi, hdrs = _find_header_grid(grid)
        if not hdrs or _is_extra_sheet(hdrs):
            continue
        fields = set(_colmap(hdrs).values())
        if "code" not in fields and "name" not in fields:
            continue
        if not headers:
            headers = hdrs
        _, recs = _records_from_grid(grid)
        all_recs.extend(recs)
    return headers, all_recs


def _stage_equipment_records(db, batch, records):
    for idx, rec in enumerate(records, start=2):
        raw = {"_sheet": "equipment"}
        raw.update(rec)
        db.add(ImportBatchRow(batch_id=batch.id, row_number=idx, raw=raw,
                              is_valid=True, staging_status="new"))


def _stage_extra_sheets(db, batch, wb):
    """Sheets 2–6 (§6B template): specs, structure, parts, PM plans and
    maintenance history — each staged independently; an error in one sheet
    never blocks the others."""
    specs = _sheet_by_hint(wb, ("مشخصات", "spec"))
    structure = _sheet_by_hint(wb, ("ساختار", "structure", "زیرسیستم"))
    parts = _sheet_by_hint(wb, ("قطعات", "part"))
    plans = _find_sheet(wb, ("برنامه", "نگهداری", "pm"))
    history = _find_sheet(wb, ("سابقه", "تاریخچه", "تعمیرات", "history"))

    def rows_of(ws, keys):
        if ws is None:
            return []
        out = []
        data = list(ws.iter_rows(values_only=True))
        if not data:
            return []
        for row in data[1:]:
            cells = [_norm(c) for c in row]
            if not any(cells):
                continue
            out.append({k: cells[i] if i < len(cells) else ""
                        for i, k in enumerate(keys)})
        return out

    def mapped_rows(ws, aliases, order):
        """Header-aware row extraction with positional fallback."""
        if ws is None:
            return []
        data = list(ws.iter_rows(values_only=True))
        if len(data) < 2:
            return []
        header = [_norm(c).lower() for c in data[0]]
        colmap = {}
        for i, hh in enumerate(header):
            if hh in aliases:
                colmap[i] = aliases[hh]
        out = []
        for row in data[1:]:
            cells = [_norm(c) for c in row]
            if not any(cells):
                continue
            rec = {}
            if colmap:
                for i, f in colmap.items():
                    if i < len(cells):
                        rec[f] = cells[i]
            else:  # positional fallback (template order)
                for i, k in enumerate(order):
                    rec[k] = cells[i] if i < len(cells) else ""
            out.append(rec)
        return out

    PM_ALIASES = {
        "کد تجهیز": "code", "code": "code",
        "عنوان": "title", "عنوان فعالیت": "title", "فعالیت": "title",
        "شرح فعالیت": "title", "title": "title",
        "نوع فعالیت": "activity_type", "نوع": "activity_type",
        "activity_type": "activity_type",
        "تناوب": "interval", "دوره تناوب": "interval", "دوره": "interval",
        "interval": "interval",
        "مجری": "performer", "مسئول": "performer", "performer": "performer",
        "مدت": "duration", "مدت (دقیقه)": "duration", "زمان استاندارد": "duration",
        "duration": "duration",
        "آخرین اجرا": "last_exec", "آخرین انجام": "last_exec",
        "last_exec": "last_exec",
    }
    HIST_ALIASES = {
        "کد تجهیز": "code", "code": "code", "equipmentcode": "code",
        "تاریخ": "date", "تاریخ انجام": "date", "date": "date",
        "day failuer": "date", "day failure": "date", "dateadd": "date",  # Access
        "تاریخ خرابی": "date",
        "نوع کار": "work_type", "نوع": "work_type", "work_type": "work_type",
        "عنوان": "title", "شرح": "title", "عنوان کار": "title", "title": "title",
        "discription": "title", "description": "title",  # Access (با غلط تایپی منبع)
        "شرح خرابی": "title",
        "اقدام تعمیراتی": "repair", "repair": "repair", "اقدام": "repair",
        "تکنسین": "technician", "technician": "technician", "مجری": "technician",
        "مدت": "duration", "مدت (دقیقه)": "duration", "duration": "duration",
        "هزینه": "cost", "هزینه (ریال)": "cost", "cost": "cost",
    }
    PART_ALIASES = {
        "کد تجهیز": "code", "code": "code", "equipmentcode": "code",
        "کد قطعه": "part_number", "part number": "part_number",
        "part_number": "part_number", "equipment spar parts": "part_number",  # Access
        "نام قطعه": "name", "spare parts": "name", "نام": "name", "name": "name",
        "تعداد": "qty", "موجودی": "qty", "موجودی فعلی": "qty", "qty": "qty",
        "حد موجودی": "min_qty", "حداقل موجودی": "min_qty",
        "inventoryminimum": "min_qty",
        "مقدار سفارش": "order_qty", "سفارش": "order_qty", "spareorder": "order_qty",
        "تأمین‌کننده": "supplier", "تامین‌کننده": "supplier",
        "تامین کننده": "supplier", "supplier": "supplier",
        "درجه اهمیت": "criticality", "criticality": "criticality",
    }

    specs_rows = rows_of(specs, ["code", "spec_name", "value", "unit"])
    structure_rows = rows_of(structure, ["code", "level", "parent_code", "name",
                                         "manufacturer", "model", "serial"])
    parts_rows = mapped_rows(parts, PART_ALIASES,
                             ["code", "part_number", "name", "qty",
                              "min_qty", "order_qty", "supplier", "criticality"])
    plans_rows = mapped_rows(plans, PM_ALIASES,
                             ["code", "title", "activity_type", "interval",
                              "performer", "duration", "last_exec"])
    history_rows = mapped_rows(history, HIST_ALIASES,
                               ["code", "date", "work_type", "title",
                                "technician", "duration", "cost"])

    n = 10_000
    for kind, rows in (("specs", specs_rows), ("structure", structure_rows),
                       ("parts", parts_rows), ("pm", plans_rows),
                       ("history", history_rows)):
        for r in rows:
            n += 1
            db.add(ImportBatchRow(batch_id=batch.id, row_number=n,
                                  raw={"_sheet": kind, **r}, is_valid=True,
                                  staging_status="new"))
    batch.summary = {"extra": {"specs": len(specs_rows), "structure": len(structure_rows),
                               "parts": len(parts_rows), "pm": len(plans_rows),
                               "history": len(history_rows)}}


def _find_sheet(wb, hints):
    for ws in wb.worksheets:
        title = (ws.title or "").strip().lower()
        if any(h in title for h in hints):
            return ws
    return None


# ---------------------------------------------------------------------------
# Mapping confirmation
# ---------------------------------------------------------------------------

class MappingIn(BaseModel):
    mapping: dict[str, str]  # {"0": "code", "1": "ignore", ...}


@router.post("/{batch_id}/mapping")
def confirm_mapping(batch_id: int, body: MappingIn, request: Request,
                    user: User = Depends(require("bulk_charge.charge")),
                    db: Session = Depends(get_db)):
    batch = db.get(ImportBatch, batch_id)
    if not batch or batch.entity_type != "bulk_charge" or batch.status != "pending":
        raise HTTPException(status_code=404, detail="بسته شارژ معتبر نیست")
    columns = []
    for idx_str, field in body.mapping.items():
        columns.append({"index": int(idx_str),
                        "field": None if field in ("ignore", "", None) else field})
    batch.mapping = {"columns": columns,
                     "header_row": (batch.mapping or {}).get("header_row", 0)}
    # re-stage values according to the confirmed mapping
    header_idx = batch.mapping["header_row"]
    audit.record(db, user_id=user.id, action="bulkcharge.mapping_confirmed",
                 entity_type="import_batch", entity_id=batch.id,
                 new=batch.mapping, request=request)
    db.commit()
    return {"ok": True, "mapping": columns, "note": f"header_row={header_idx}"}


# ---------------------------------------------------------------------------
# Preview / diff
# ---------------------------------------------------------------------------

@router.get("/{batch_id}/preview")
def preview(batch_id: int, user: User = Depends(require("bulk_charge.charge")),
            db: Session = Depends(get_db)):
    batch = db.get(ImportBatch, batch_id)
    if not batch or batch.entity_type != "bulk_charge":
        raise HTTPException(status_code=404, detail="بسته شارژ معتبر نیست")

    existing = db.query(Equipment).filter(Equipment.deleted_at.is_(None)).all()
    ex_by_code = {e.code: e for e in existing}
    factories = {f.name.strip().lower(): f.name for f in db.query(Factory).all()} | \
                {f.code.strip().lower(): f.name for f in db.query(Factory).all()}
    cats = {c.name.strip().lower(): c.name for c in db.query(EquipmentCategory).all()} | \
           {c.code.strip().lower(): c.name for c in db.query(EquipmentCategory).all()}

    counts = {"new": 0, "update": 0, "conflict": 0, "rejected": 0}
    rows_out = []
    rows = db.query(ImportBatchRow).filter(ImportBatchRow.batch_id == batch.id).all()
    # codes staged in THIS batch (pm/history rows may reference them)
    staged_codes = {
        (r.raw or {}).get("code") for r in rows
        if (r.raw or {}).get("_sheet") in (None, "equipment") and (r.raw or {}).get("code")
    }
    for r in sorted(rows, key=lambda x: x.row_number):
        raw = dict(r.raw or {})  # copy! JSON columns need new object to track
        if raw.get("_sheet") not in (None, "equipment"):
            continue
        errs = list(r.errors or [])
        status = r.staging_status or "new"
        matched_id = None

        if status != "rejected":
            code = raw.get("code", "")
            name = raw.get("name", "")
            factory = raw.get("factory", "")
            category = raw.get("category", "")

            # §7: decode the code first — factory/category helper signals
            if code:
                dec = decode_code(db, code)
                if dec.get("unknown_prefix"):
                    raw["_unknown_prefix"] = True
                else:
                    if dec.get("factory_name") and not factory:
                        raw["_inferred_factory"] = dec["factory_name"]
                        factory = dec["factory_name"]
                    elif dec.get("factory_name") and factory and \
                            factory.strip() != dec["factory_name"]:
                        # §7: column factory disagrees with code prefix → flag,
                        # never auto-fix
                        status = "conflict"
                        errs.append(f"کارخانه ستون «{factory}» با پیشوند کد "
                                    f"«{dec['factory_name']}» ناهماهنگ است — بدون اصلاح خودکار")
                    if dec.get("area_name") and not category:
                        raw["_inferred_category"] = dec["area_name"]
                        raw["_inferred_area"] = dec["area_code"]
                        category = dec["area_name"]

            eff = dict(raw)
            eff["factory"] = factory
            eff["category"] = category
            missing = [f for f in REQUIRED if not eff.get(f)]
            if raw.get("_unknown_prefix"):
                status = "conflict"
                errs.append("پیشوند ناشناخته در کد — نیاز به تعریف توسط Admin")
            elif missing:
                errs.append("فیلدهای اجباری ناقص: " + "، ".join(missing))
                status = "rejected"
            else:
                if code in ex_by_code:
                    status = "update"
                    matched_id = ex_by_code[code].id
                else:
                    for e in existing:
                        score = max(
                            _similarity(name, e.name),
                            (_similarity(name, e.name) +
                             _similarity(factory, e.factory.name if e.factory else "")) / 2,
                        )
                        serial_same = (raw.get("serial_number") and e.serial_number and
                                       raw.get("serial_number") == e.serial_number)
                        if score >= 0.85 or serial_same:
                            status = "conflict"
                            matched_id = e.id
                            errs.append(f"احتمال تکراری با {e.code} — {e.name} "
                                        f"(شباهت {int(score * 100)}٪)؛ تصمیم با کاربر")
                            break
                    fkey = factory.strip().lower()
                    if fkey and fkey not in factories:
                        raw["_new_factory"] = factory
                        errs.append(f"کارخانه «{factory}» جدید است — با Commit ایجاد می‌شود")
                    ckey = category.strip().lower()
                    if ckey and ckey not in cats:
                        raw["_new_category"] = category
                        errs.append(f"دسته «{category}» جدید است — با Commit ایجاد می‌شود")
                    unmapped = [u for u in raw.get("_unmapped", []) if not u.get("guess")]
                    if len(unmapped) > 2:
                        status = "conflict"
                        errs.append("نیاز به بازبینی دستی: سلول‌های غیرقابل‌تشخیص زیاد است")

        r.staging_status = status
        r.matched_equipment_id = matched_id
        r.errors = errs
        r.raw = raw  # persist inferred values back to the JSON column
        counts[status] = counts.get(status, 0) + 1
        rows_out.append({
            "row_id": r.id, "row_number": r.row_number, "status": status,
            "code": raw.get("code"), "name": raw.get("name"),
            "factory": raw.get("factory") or raw.get("_inferred_factory"),
            "category": raw.get("category"), "errors": errs,
            "matched_equipment_id": matched_id,
            "resolution": r.resolution,
        })

    # pm / history sheets: equipment reference must exist (DB or this batch)
    for r in rows:
        raw = dict(r.raw or {})
        if raw.get("_sheet") not in ("pm", "history"):
            continue
        sheet_fa = "برنامه نت" if raw.get("_sheet") == "pm" else "سابقه تعمیرات"
        errs = list(r.errors or [])
        status = r.staging_status or "new"
        if status != "rejected" and raw.get("_checked") is not True:
            code = raw.get("code", "")
            if not raw.get("title") or not code:
                errs.append(f"{sheet_fa}: کد تجهیز و عنوان الزامی است")
                status = "rejected"
            elif code not in ex_by_code and code not in staged_codes:
                errs.append(f"{sheet_fa}: تجهیز با کد {code} یافت نشد")
                status = "rejected"
            elif raw.get("_sheet") == "pm":
                icode, _ = _interval_from(db, raw.get("interval") or "")
                if icode is None:
                    errs.append(f"{sheet_fa}: تناوب «{raw.get('interval')}» نامعتبر است")
                    status = "conflict"
                d = raw.get("last_exec") or ""
                if d and _parse_flexible_date(d) is None:
                    errs.append(f"{sheet_fa}: تاریخ «{d}» قابل خواندن نیست")
                    status = "conflict"
            elif raw.get("_sheet") == "history":
                d = raw.get("date") or ""
                if d and _parse_flexible_date(d) is None:
                    errs.append(f"{sheet_fa}: تاریخ «{d}» قابل خواندن نیست")
                    status = "conflict"
            raw["_checked"] = True
            r.raw = raw
        r.staging_status = status
        r.errors = errs
        if status != "new":
            counts[status] = counts.get(status, 0) + 1
        rows_out.append({
            "row_id": r.id, "row_number": r.row_number, "status": status,
            "code": raw.get("code"), "name": raw.get("title"),
            "factory": None, "category": sheet_fa, "errors": errs,
            "matched_equipment_id": None, "resolution": r.resolution,
        })

    extra = (batch.summary or {}).get("extra", {})
    db.commit()
    return {"batch_id": batch.id, "counts": counts, "rows": rows_out,
            "extra_sheets": extra,
            "note": "Commit فقط رکوردهای بدون Conflict را منتقل می‌کند"}


class RowEditIn(BaseModel):
    field: str
    value: str


@router.post("/{batch_id}/rows/{row_id}")
def edit_row(batch_id: int, row_id: int, body: RowEditIn, request: Request,
             user: User = Depends(require("bulk_charge.charge")),
             db: Session = Depends(get_db)):
    batch = db.get(ImportBatch, batch_id)
    if not batch or batch.status != "pending":
        raise HTTPException(status_code=400, detail="بسته در وضعیت ویرایش نیست")
    r = db.get(ImportBatchRow, row_id)
    if r is None or r.batch_id != batch.id:
        raise HTTPException(status_code=404, detail="ردیف یافت نشد")
    raw = dict(r.raw or {})
    raw[body.field] = body.value
    raw.pop("_unmapped", None)
    r.raw = raw
    if r.staging_status == "rejected":
        r.staging_status = "new"
        r.errors = []
    audit.record(db, user_id=user.id, action="bulkcharge.row_edited",
                 entity_type="import_batch_row", entity_id=r.id,
                 new={"field": body.field, "value": body.value}, request=request)
    db.commit()
    return {"ok": True}


class ResolveIn(BaseModel):
    action: str  # create_new | merge | reject


@router.post("/{batch_id}/rows/{row_id}/resolve")
def resolve_row(batch_id: int, row_id: int, body: ResolveIn, request: Request,
                user: User = Depends(require("bulk_charge.approve")),
                db: Session = Depends(get_db)):
    r = db.get(ImportBatchRow, row_id)
    if r is None or r.batch_id != batch_id:
        raise HTTPException(status_code=404, detail="ردیف یافت نشد")
    if body.action not in ("create_new", "merge", "reject"):
        raise HTTPException(status_code=400, detail="کنش نامعتبر است")
    if body.action == "reject":
        r.staging_status = "rejected"
        r.resolution = "رد توسط کاربر"
    elif body.action == "merge" and r.matched_equipment_id:
        r.staging_status = "update"
        r.resolution = "ادغام با رکورد موجود"
    else:
        r.staging_status = "new"
        r.matched_equipment_id = None
        r.resolution = "ایجاد رکورد جدید"
    audit.record(db, user_id=user.id, action="bulkcharge.row_resolved",
                 entity_type="import_batch_row", entity_id=r.id,
                 new={"action": body.action}, request=request)
    db.commit()
    return {"ok": True, "status": r.staging_status}


# ---------------------------------------------------------------------------
# Commit (§6B) — clean rows only
# ---------------------------------------------------------------------------

@router.post("/{batch_id}/commit")
def commit(batch_id: int, request: Request,
           user: User = Depends(require("bulk_charge.approve")),
           db: Session = Depends(get_db)):
    batch = db.get(ImportBatch, batch_id)
    if not batch or batch.entity_type != "bulk_charge" or batch.status != "pending":
        raise HTTPException(status_code=400, detail="بسته قابل Commit نیست")

    factories = {f.name.strip().lower(): f for f in db.query(Factory).all()} | \
                {f.code.strip().lower(): f for f in db.query(Factory).all()}
    cats = {c.name.strip().lower(): c for c in db.query(EquipmentCategory).all()} | \
            {c.code.strip().lower(): c for c in db.query(EquipmentCategory).all()}

    rows = db.query(ImportBatchRow).filter(ImportBatchRow.batch_id == batch.id).all()
    eq_rows = [r for r in rows if (r.raw or {}).get("_sheet") in (None, "equipment")]
    created = updated = skipped = 0

    for r in eq_rows:
        if r.staging_status in ("conflict", "rejected"):
            skipped += 1
            continue
        raw = r.raw or {}
        if not raw.get("code") or not raw.get("name"):
            skipped += 1
            continue

        fkey = (raw.get("factory") or raw.get("_inferred_factory", "")).strip().lower()
        factory = factories.get(fkey)
        if factory is None and (raw.get("factory") or raw.get("_inferred_factory")):
            fname = raw.get("factory") or raw.get("_inferred_factory")
            factory = Factory(code=f"BCH-{len(factories)+1:03d}",
                              name=fname, created_by=user.id)
            db.add(factory); db.flush()
            factories[factory.name.strip().lower()] = factory
        ckey = (raw.get("category") or raw.get("_inferred_category", "")).strip().lower()
        category = cats.get(ckey)
        if category is None and (raw.get("category") or raw.get("_inferred_category")):
            cname = raw.get("category") or raw.get("_inferred_category")
            area = raw.get("_inferred_area")
            category = EquipmentCategory(
                code=(f"CAT-{area}" if area else f"BCHC-{len(cats)+1:03d}"),
                name=cname, created_by=user.id)
            db.add(category); db.flush()
            cats[category.name.strip().lower()] = category

        crit, crit_score = resolve_criticality(raw.get("criticality"))
        status = STATUS_ALIASES.get((raw.get("status") or "active").strip().lower(), "active")
        year = None
        if raw.get("year"):
            try:
                year = int(float(raw["year"]))
            except ValueError:
                year = None

        existing = db.query(Equipment).filter(Equipment.code == raw["code"]).one_or_none()
        if existing and r.staging_status == "update":
            old = {"name": existing.name, "manufacturer": existing.manufacturer,
                   "model": existing.model, "serial_number": existing.serial_number,
                   "year": existing.year, "criticality": existing.criticality,
                   "status": existing.status, "version": existing.version}
            r.raw = {**raw, "_old": old}
            existing.name = raw.get("name") or existing.name
            existing.manufacturer = raw.get("manufacturer") or existing.manufacturer
            existing.model = raw.get("model") or existing.model
            existing.serial_number = raw.get("serial_number") or existing.serial_number
            if year:
                existing.year = year
            existing.criticality = crit
            if crit_score is not None:
                existing.criticality_score = crit_score
            existing.status = status
            existing.hall = raw.get("hall") or existing.hall
            existing.dept = raw.get("dept") or existing.dept
            existing.line = raw.get("line") or raw.get("product_line") or existing.line or raw.get("_section")
            existing.position = raw.get("location") or existing.position
            _dyn = dict(existing.dynamic_fields or {})
            _dyn.update(extra_dyn_fields(raw))
            existing.dynamic_fields = _dyn or None
            existing.version += 1
            existing.updated_by = user.id
            existing.updated_at = utcnow()
            r.created_equipment_id = existing.id
            updated += 1
        else:
            dyn = extra_dyn_fields(raw)
            e = Equipment(
                code=raw["code"], name=raw["name"], level="equipment",
                factory_id=factory.id if factory else None,
                category_id=category.id if category else None,
                manufacturer=raw.get("manufacturer") or None,
                model=raw.get("model") or None,
                serial_number=raw.get("serial_number") or None,
                year=year, criticality=crit, status=status,
                criticality_score=crit_score,
                hall=raw.get("hall") or None, dept=raw.get("dept") or None,
                line=raw.get("line") or raw.get("_section") or None,
                position=raw.get("location") or None,
                component_type=raw.get("component_type") or None,
                dynamic_fields=dyn or None,
                created_by=user.id,
            )
            specs = {}
            for k in ("power", "country"):
                if raw.get(k):
                    specs[{"power": "توان", "country": "کشور سازنده"}[k]] = raw[k]
            if raw.get("capacity"):
                unit = raw.get("capacity_unit") or ""
                specs["ظرفیت"] = f"{raw['capacity']} {unit}".strip()
            if specs:
                e.technical_specs = specs
            db.add(e); db.flush()
            r.created_equipment_id = e.id
            created += 1
        db.flush()

    # sheet 2: long-format specs → technical_specs
    specs_applied = 0
    for r in rows:
        raw = r.raw or {}
        if raw.get("_sheet") != "specs":
            continue
        eq = db.query(Equipment).filter(Equipment.code == raw.get("code")).one_or_none()
        if eq is None or not raw.get("spec_name"):
            continue
        ts = dict(eq.technical_specs or {})
        key = raw["spec_name"] + (f" ({raw['unit']})" if raw.get("unit") else "")
        ts[key] = raw.get("value", "")
        eq.technical_specs = ts
        specs_applied += 1

    # sheet 3: structure (subsystem/component/subcomponent)
    LEVEL_FA = {"زیرسیستم": "subsystem", "subsystem": "subsystem",
                "قطعه": "component", "component": "component", "جزء": "component",
                "زیرقطعه": "subcomponent", "subcomponent": "subcomponent"}
    structure_created = 0
    for r in rows:
        raw = r.raw or {}
        if raw.get("_sheet") != "structure":
            continue
        parent_code = raw.get("parent_code") or raw.get("code")
        parent = db.query(Equipment).filter(Equipment.code == parent_code).one_or_none()
        if parent is None or not raw.get("name"):
            continue
        level = LEVEL_FA.get((raw.get("level") or "").strip().lower())
        if level is None:
            continue
        code = raw.get("code") if raw.get("level") else f"{parent.code}-S{structure_created+1}"
        db.add(Equipment(
            code=f"{parent.code}-{level[0].upper()}{structure_created+1}",
            name=raw["name"], level=level, parent_id=parent.id,
            factory_id=parent.factory_id, category_id=parent.category_id,
            manufacturer=raw.get("manufacturer") or None,
            model=raw.get("model") or None,
            serial_number=raw.get("serial") or None,
            created_by=user.id,
        ))
        structure_created += 1

    # sheet 4: parts
    parts_created = 0
    for r in rows:
        raw = r.raw or {}
        if raw.get("_sheet") != "parts":
            continue
        eq = db.query(Equipment).filter(Equipment.code == raw.get("code")).one_or_none()
        if eq is None or not raw.get("name"):
            continue
        if raw.get("part_number") and db.query(Part).filter(
                Part.code == raw.get("part_number")).one_or_none():
            continue
        from ..modules.parts import _resolve_supplier
        db.add(Part(code=raw.get("part_number") or f"BCHP-{parts_created+1}",
                    name=raw["name"],
                    stock_qty=float(raw["qty"]) if _is_num(raw.get("qty")) else 0,
                    min_qty=float(raw["min_qty"]) if _is_num(raw.get("min_qty")) else 0,
                    order_qty=float(raw["order_qty"]) if _is_num(raw.get("order_qty")) else None,
                    supplier=raw.get("supplier") or None,
                    supplier_id=_resolve_supplier(db, raw.get("supplier"), None),
                    criticality=CRIT_ALIASES.get((raw.get("criticality") or "medium").lower(), "medium"),
                    equipment_id=eq.id, import_batch_id=batch.id, created_by=user.id))
        parts_created += 1

    # sheet 5: maintenance plans (PM programs)
    from ..models import MaintenancePlan
    from datetime import timedelta

    plans_created = 0
    plan_ids: list[int] = []
    for r in rows:
        raw = r.raw or {}
        if raw.get("_sheet") != "pm":
            continue
        eq = db.query(Equipment).filter(Equipment.code == raw.get("code")).one_or_none()
        if eq is None or not raw.get("title"):
            continue
        icode, idays = _interval_from(db, raw.get("interval") or "")
        if icode is None:
            continue
        last = _parse_flexible_date(raw.get("last_exec") or "")
        plan = MaintenancePlan(
            equipment_id=eq.id, work_class="pm", work_title=raw["title"],
            activity_type=_activity_from(db, raw.get("activity_type") or ""),
            interval_code=icode, interval_days=idays,
            performer=raw.get("performer") or None,
            duration_minutes=int(float(raw["duration"])) if _is_num(raw.get("duration")) else None,
            last_execution=last,
            next_due=(last + timedelta(days=idays)) if last else None,
            created_by=user.id,
        )
        db.add(plan); db.flush()
        plan_ids.append(plan.id)
        plans_created += 1

    # sheet 6: maintenance history → دستورکار بسته‌شده (§4.3 سند بارگذاری نهایی)
    from ..models import MaintenanceHistory, WorkOrder

    history_created = 0
    history_ids: list[int] = []
    wo_ids: list[int] = []
    wo_seq: dict[int, int] = {}
    for r in rows:
        raw = r.raw or {}
        if raw.get("_sheet") != "history":
            continue
        eq = db.query(Equipment).filter(Equipment.code == raw.get("code")).one_or_none()
        if eq is None or not raw.get("title"):
            continue
        if eq.id not in wo_seq:  # ادامه شماره‌گذاری از دستورکارهای موجود
            wo_seq[eq.id] = db.query(WorkOrder).filter(
                WorkOrder.code.like(f"{eq.code}-WO-%")).count()
        at = _parse_flexible_date(raw.get("date") or "")
        title = raw["title"]
        if raw.get("repair"):
            title = f"{title} — {raw['repair']}"

        # --- تخصیص تکنسین طبق قانون نیروی انسانی (بخش ۵ سند) ---
        tech = None
        if raw.get("technician"):
            tech = db.query(User).filter(User.username == raw["technician"]).one_or_none() \
                or db.query(User).filter(User.full_name == raw["technician"]).one_or_none()
        if tech is None:
            tech = _assign_history_technician(db, eq, title, at)

        # --- دستورکار بسته‌شده با کد {تجهیز}-WO-{شماره} ---
        wo_seq[eq.id] = wo_seq.get(eq.id, 0) + 1
        wo = WorkOrder(
            code=f"{eq.code}-WO-{wo_seq[eq.id]:02d}",
            title=title[:190],
            description=(f"هزینه ثبت‌شده: {raw['cost']} ریال" if raw.get("cost") else None),
            equipment_id=eq.id, status="closed",
            work_class="cm", priority="normal",
            assigned_to=tech.id if tech else None,
            completed_at=at, created_by=user.id,
        )
        db.add(wo); db.flush()
        wo_ids.append(wo.id)

        h = MaintenanceHistory(
            equipment_id=eq.id, work_order_id=wo.id,
            work_type=raw.get("work_type") or "تعمیر",
            title=title[:190],
            description=(f"هزینه ثبت‌شده: {raw['cost']} ریال" if raw.get("cost") else None),
            technician_id=tech.id if tech else None,
            finished_at=at, started_at=at,
            duration_minutes=int(float(raw["duration"])) if _is_num(raw.get("duration")) else None,
        )
        db.add(h); db.flush()
        history_ids.append(h.id)
        history_created += 1

    batch.status = "confirmed"
    batch.confirmed_at = utcnow()
    batch.valid_rows = created + updated
    batch.error_rows = skipped
    batch.summary = {**(batch.summary or {}),
                     "created": created, "updated": updated, "skipped": skipped,
                     "specs_applied": specs_applied,
                     "structure_created": structure_created,
                     "parts_created": parts_created,
                     "plans_created": plans_created,
                     "history_created": history_created,
                     "created_plan_ids": plan_ids,
                     "created_history_ids": history_ids,
                     "created_wo_ids": wo_ids}
    audit.record(db, user_id=user.id, action="bulkcharge.committed",
                 entity_type="import_batch", entity_id=batch.id,
                 new=batch.summary, request=request)
    db.commit()
    bus.publish("equipment.bulk_imported", {"batch_id": batch.id, "created": created})
    return {"ok": True, **batch.summary}


# ---------------------------------------------------------------------------
# Rollback (§6B) — batch level, with modification guard
# ---------------------------------------------------------------------------

@router.post("/{batch_id}/rollback")
def rollback(batch_id: int, request: Request,
             user: User = Depends(require("bulk_charge.rollback")),
             db: Session = Depends(get_db)):
    batch = db.get(ImportBatch, batch_id)
    if not batch or batch.entity_type != "bulk_charge" or batch.status != "confirmed":
        raise HTTPException(status_code=400, detail="فقط بسته Commit‌شده قابل بازگردانی است")

    from ..models import FileObject, MaintenancePlan, MaintenanceHistory
    from sqlalchemy import delete as _sqdelete

    # Phase A — remove batch-created dependents FIRST so equipment FKs
    # are free when we purge the created equipment rows.
    plans_removed = plans_conflicts = 0
    for pid in (batch.summary or {}).get("created_plan_ids", []):
        p = db.get(MaintenancePlan, pid)
        if p is None:
            continue
        if p.version != 1:  # edited after commit → keep, report conflict
            plans_conflicts += 1
            continue
        db.delete(p)
        plans_removed += 1
    db.flush()
    history_removed = 0
    for hid in (batch.summary or {}).get("created_history_ids", []):
        hrow = db.get(MaintenanceHistory, hid)
        if hrow is not None:
            db.delete(hrow)
            history_removed += 1
    db.flush()
    from ..models import WorkOrder
    wo_removed = 0
    for wid in (batch.summary or {}).get("created_wo_ids", []):
        wrow = db.get(WorkOrder, wid)
        if wrow is not None:
            db.delete(wrow)
            wo_removed += 1
    db.flush()
    parts_removed = db.query(Part).filter(Part.import_batch_id == batch.id).delete()
    db.flush()

    # Phase B — equipment rows: restore updates / purge creations.
    def _purge_equipment(e):
        """Delete an equipment created by this batch along with rows that
        reference it (children built from the structure sheet, parts, plans,
        work orders, history, files) so the FK graph never blocks the batch
        rollback."""
        for child in db.query(Equipment).filter(Equipment.parent_id == e.id).all():
            _purge_equipment(child)
        db.query(MaintenancePlan).filter(
            MaintenancePlan.equipment_id == e.id).delete(synchronize_session=False)
        db.query(MaintenancePlan).filter(
            MaintenancePlan.target_id == e.id).update(
            {"target_id": None}, synchronize_session=False)
        db.query(MaintenanceHistory).filter(
            MaintenanceHistory.equipment_id == e.id).delete(synchronize_session=False)
        db.query(WorkOrder).filter(
            WorkOrder.equipment_id == e.id).delete(synchronize_session=False)
        for f in db.query(FileObject).filter(
                FileObject.entity_type == "equipment",
                FileObject.entity_id == e.id).all():
            db.delete(f)
        db.query(Part).filter(Part.equipment_id == e.id).delete(synchronize_session=False)
        db.delete(e)
        db.flush()  # enforce child→parent delete order for self-referential FKs

    removed = restored = conflicts = 0
    rows = db.query(ImportBatchRow).filter(ImportBatchRow.batch_id == batch.id).all()
    for r in rows:
        if not r.created_equipment_id:
            continue
        e = db.get(Equipment, r.created_equipment_id)
        if e is None:
            continue
        old = (r.raw or {}).get("_old")
        if old:  # was an update → restore snapshot if untouched since
            # §6B guard: any direct change after commit bumped the version.
            if e.version != old["version"] + 1:
                conflicts += 1
                continue
            e.name = old.get("name", e.name)
            e.manufacturer = old.get("manufacturer")
            e.model = old.get("model")
            e.serial_number = old.get("serial_number")
            e.year = old.get("year")
            e.criticality = old.get("criticality", e.criticality)
            e.status = old.get("status", e.status)
            e.version += 1
            restored += 1
        else:  # was a creation → remove (with dependent rows) if untouched
            # created rows start at version 1; any later edit bumps it (§35)
            if e.version != 1:
                conflicts += 1
                continue
            r.created_equipment_id = None  # release the staging FK first
            db.flush()
            _purge_equipment(e)
            removed += 1
            continue
        r.created_equipment_id = None

    # parts/plans/history already handled in Phase A above.
    batch.status = "rolled_back"
    summary = {"removed": removed, "restored": restored,
               "conflicts": conflicts + plans_conflicts,
               "parts_removed": parts_removed,
               "plans_removed": plans_removed,
               "history_removed": history_removed}
    audit.record(db, user_id=user.id, action="bulkcharge.rolled_back",
                 entity_type="import_batch", entity_id=batch.id,
                 new=summary, request=request)
    db.commit()
    return {"ok": True, **summary,
            "warning": ("برخی رکوردها پس از Commit تغییر کرده‌اند و بازگردانی نشدند"
                        if conflicts else None)}
