"""Equipment Bulk Data Charge (MODULE EQUIPMENT — BASPAR).

Flow (Master-prompt §23 pattern applied to equipment legacy data):
    Upload Excel/Markdown → Preview + Validation → Confirm → (optionally) Rollback

Designed for messy legacy exports: Persian/English headers, auto-creation
of missing factories/categories, parent rows defined inside the same file.

Markdown import (user delivers equipment files as ``.md``):
    * Markdown tables  → header row mapped through HEADER_ALIASES
    * Heading sections → each ``## Heading`` starts a record, ``- key: value``
      bullets fill the fields
"""
from __future__ import annotations

import io
import re

from fastapi import APIRouter, Depends, File, HTTPException, Request, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from pydantic import BaseModel, Field
from sqlalchemy.orm import Session

from .. import audit
from ..db import get_db, utcnow
from ..events import bus
from ..models import Equipment, EquipmentCategory, Factory, ImportBatch, ImportBatchRow, User
from ..rbac import require

router = APIRouter(prefix="/equipment/bulk-import", tags=["bulk-import"])

# Persian/English header aliases → canonical field
HEADER_ALIASES: dict[str, str] = {
    "code": "code", "کد تجهیز": "code", "کد": "code",
    "name": "name", "نام تجهیز": "name", "نام": "name",
    "level": "level", "سطح": "level", "نوع": "level",
    "factory": "factory", "کارخانه": "factory",
    "category": "category", "دسته‌بندی": "category", "کلاس تجهیز": "category",
    "parent_code": "parent_code", "کد والد": "parent_code", "والد": "parent_code",
    "location": "location", "محل نصب": "location", "موقعیت": "location",
    "manufacturer": "manufacturer", "سازنده": "manufacturer",
    "model": "model", "مدل": "model",
    "serial_number": "serial_number", "شماره سریال": "serial_number", "سریال": "serial_number",
    "year": "year", "سال ساخت": "year", "سال": "year",
    "criticality": "criticality", "درجه اهمیت": "criticality", "بحرانیت": "criticality",
    "اهمیت": "criticality",
    "status": "status", "وضعیت": "status",
    "component_type": "component_type", "نوع قطعه": "component_type",
    "نوع تجهیز": "component_type", "کلاس قطعه": "component_type",
    "hall": "hall", "سالن": "hall",
    "dept": "dept", "بخش": "dept", "دپارتمان": "dept",
    "line": "line", "خط": "line", "خط تولید": "line",
    "position": "position", "پست": "position", "محل دقیق": "position",
    "location_notes": "location_notes", "توضیحات محل": "location_notes",
    "notes": "notes", "توضیحات": "notes", "یادداشت": "notes",
}

LEVEL_MAP = {
    "equipment": "equipment", "تجهیز": "equipment", "تجهیز اصلی": "equipment",
    "subsystem": "subsystem", "زیرسیستم": "subsystem",
    "component": "component", "جزء": "component", "قطعه": "component",
    "subcomponent": "subcomponent", "زیرقطعه": "subcomponent", "زیر جزء": "subcomponent",
    "زیرجزء": "subcomponent",
}

CRITICALITY_MAP = {
    "low": "low", "کم": "low", "پایین": "low",
    "medium": "medium", "متوسط": "medium",
    "high": "high", "زیاد": "high", "بالا": "high",
    "critical": "critical", "بحرانی": "critical",
}

STATUS_MAP = {
    "active": "active", "فعال": "active",
    "inactive": "inactive", "غیرفعال": "inactive",
    "under_maintenance": "under_maintenance", "در دست تعمیر": "under_maintenance",
    "scrapped": "scrapped", "اسقاط": "scrapped", "از رده خارج": "scrapped",
}

TEMPLATE_COLUMNS = [
    ("کد تجهیز", "code", "شناسه یکتا (الزامی) — مثال: EQ-1001"),
    ("نام تجهیز", "name", "نام تجهیز (الزامی)"),
    ("سطح", "level", "تجهیز / زیرسیستم / جزء / زیرقطعه"),
    ("کارخانه", "factory", "نام یا کد کارخانه"),
    ("دسته‌بندی", "category", "نام یا کد دسته‌بندی"),
    ("کد والد", "parent_code", "کد تجهیز والد (برای زیرسیستم/جزء/زیرقطعه)"),
    ("محل نصب", "location", "محل فیزیکی نصب"),
    ("سالن", "hall", "سالن تولید (اختیاری)"),
    ("بخش", "dept", "بخش/دپارتمان (اختیاری)"),
    ("خط", "line", "خط تولید (اختیاری)"),
    ("نوع قطعه", "component_type", "پمپ / تابلو برق / کمپرسور / …"),
    ("سازنده", "manufacturer", "شرکت سازنده"),
    ("مدل", "model", "مدل تجهیز"),
    ("شماره سریال", "serial_number", "سریال ساخت"),
    ("سال ساخت", "year", "سال میلادی — مثال: 2019"),
    ("درجه اهمیت", "criticality", "کم / متوسط / زیاد / بحرانی"),
    ("وضعیت", "status", "فعال / غیرفعال / در دست تعمیر / اسقاط"),
]


def _norm(v) -> str:
    if v is None:
        return ""
    return str(v).strip()


# ---------------------------------------------------------------------------
# Markdown / plain-text parsing
# ---------------------------------------------------------------------------

_SEPARATOR_CELL = re.compile(r"^:?-{2,}:?$")
_HEADING = re.compile(r"^#{1,6}\s+(.*)$")
_KEYVALUE = re.compile(r"^(?:[-*•]\s+)?(.+?)\s*[:：]\s*(.*)$")
_TITLE_SPLIT = re.compile(r"\s+[—–\\-]{1,2}\s+|\s*[:：]\s*")


def _clean_md(text: str) -> str:
    """Strip markdown emphasis/backticks from a cell value."""
    return re.sub(r"[*_`]+", "", text or "").strip()


def _split_md_row(line: str) -> list[str]:
    s = line.strip()
    if s.startswith("|"):
        s = s[1:]
    if s.endswith("|"):
        s = s[:-1]
    return [_clean_md(c) for c in s.split("|")]


def _is_separator_row(cells: list[str]) -> bool:
    return bool(cells) and all(_SEPARATOR_CELL.match(c or "") or c == "" for c in cells) \
        and any("-" in (c or "") for c in cells)


def _map_header(cells: list[str]) -> dict[int, str]:
    colmap: dict[int, str] = {}
    for i, cell in enumerate(cells):
        key = _norm(cell).lower()
        if key in HEADER_ALIASES:
            colmap[i] = HEADER_ALIASES[key]
    return colmap


def parse_markdown_equipment(text: str) -> list[dict]:
    """Parse Markdown (tables or heading sections) into equipment row dicts."""
    lines = (text or "").replace("\r\n", "\n").split("\n")

    rows: list[dict] = []

    # 1) Markdown table blocks -------------------------------------------
    blocks: list[list[str]] = []
    cur: list[str] = []
    for ln in lines:
        if ln.strip().startswith("|"):
            cur.append(ln)
        else:
            if cur:
                blocks.append(cur)
                cur = []
    if cur:
        blocks.append(cur)

    for block in blocks:
        grid = [_split_md_row(ln) for ln in block]
        grid = [g for g in grid if not _is_separator_row(g) and any(_norm(c) for c in g)]
        if len(grid) < 2:
            continue
        colmap = _map_header(grid[0])
        if "code" not in colmap.values() and "name" not in colmap.values():
            continue  # not an equipment table
        for data in grid[1:]:
            rec = {field: _norm(data[i]) if i < len(data) else ""
                   for i, field in colmap.items()}
            if any(rec.values()):
                rows.append(rec)

    if rows:
        return rows

    # 2) Heading-based sections -------------------------------------------
    records: list[dict] = []
    current: dict | None = None
    for ln in lines:
        s = ln.strip()
        m = _HEADING.match(s)
        if m:
            if current is not None:
                records.append(current)
            title = _clean_md(m.group(1))
            current = {"code": "", "name": "", "_title": title}
            continue
        if current is None or not s or s.startswith("|"):
            continue
        m2 = _KEYVALUE.match(s)
        if not m2:
            continue
        key = _clean_md(m2.group(1)).lower()
        value = _clean_md(m2.group(2))
        field = HEADER_ALIASES.get(key)
        if field:
            current[field] = value

    if current is not None:
        records.append(current)

    for rec in records:
        title = rec.pop("_title", "")
        if not rec.get("code") and not rec.get("name") and title:
            # derive code/name from heading like «B3P1 — پمپ شماره ۱»
            parts = [p.strip() for p in _TITLE_SPLIT.split(title) if p and p.strip()]
            if len(parts) >= 2 and re.search(r"[A-Za-z0-9]", parts[0]):
                rec["code"] = parts[0]
                rec["name"] = " ".join(parts[1:]) or title
            else:
                rec["code"] = title
                rec["name"] = title
        rec.pop("_title", None)
        if any(v for k, v in rec.items() if k != "level"):
            rows.append(rec)

    return rows


class TextImportIn(BaseModel):
    text: str = Field(min_length=1, max_length=2_000_000)
    filename: str | None = "markdown-paste.md"
    auto_create_lookups: bool = False



@router.get("/template")
def download_template(_: User = Depends(require("equipment.view"))):
    wb = Workbook()
    ws = wb.active
    ws.title = "Equipment"
    ws.append([c[0] for c in TEMPLATE_COLUMNS])
    ws.append(["EQ-1001", "کمپرسور هوای فشرده شماره ۱", "تجهیز", "کارخانه مرکزی",
               "تأسیسات", "", "سالن تولید ۱", "سالن ۱", "تأسیسات", "",
               "کمپرسور", "Atlas Copco", "GA-75", "SN-99812",
               2019, "زیاد", "فعال"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    headers = {"Content-Disposition": "attachment; filename=equipment-import-template.xlsx"}
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


def _validate_batch(db: Session, batch: ImportBatch, rows: list[dict],
                    existing_codes: set[str]) -> tuple[int, int]:
    """Fill ImportBatchRow records; returns (valid, errors)."""
    factories = db.query(Factory).all()
    categories = db.query(EquipmentCategory).all()
    f_by_key = {f.code.strip().lower(): f for f in factories} | \
               {f.name.strip().lower(): f for f in factories}
    c_by_key = {c.code.strip().lower(): c for c in categories} | \
               {c.name.strip().lower(): c for c in categories}

    file_codes: dict[str, int] = {}  # code -> row_number
    valid = errors = 0

    for idx, raw in enumerate(rows, start=2):  # row 1 = header
        errs: list[str] = []
        code = _norm(raw.get("code"))
        name = _norm(raw.get("name"))

        if not code:
            errs.append("کد تجهیز الزامی است")
        elif code in existing_codes:
            errs.append("کد تجهیز در سامانه وجود دارد")
        elif code in file_codes:
            errs.append(f"کد تجهیز در فایل تکراری است (ردیف {file_codes[code]})")
        if not name:
            errs.append("نام تجهیز الزامی است")

        level = LEVEL_MAP.get(_norm(raw.get("level")).lower(), "")
        if not level:
            errs.append("سطح نامعتبر است (تجهیز/زیرسیستم/جزء/زیرقطعه)")

        factory_key = _norm(raw.get("factory")).lower()
        category_key = _norm(raw.get("category")).lower()
        if level == "equipment":
            if not factory_key or factory_key not in f_by_key:
                if not batch.auto_create_lookups or not factory_key:
                    errs.append("کارخانه یافت نشد (ایجاد خودکار فعال نیست یا مقدار خالی است)")
            if not category_key or category_key not in c_by_key:
                if not batch.auto_create_lookups or not category_key:
                    errs.append("دسته‌بندی یافت نشد (ایجاد خودکار فعال نیست یا مقدار خالی است)")

        year_raw = _norm(raw.get("year"))
        year = None
        if year_raw:
            try:
                year = int(float(year_raw))
                if not (1800 <= year <= 2200):
                    raise ValueError
            except ValueError:
                errs.append("سال ساخت نامعتبر است")

        crit_raw = _norm(raw.get("criticality")).lower() or "medium"
        if crit_raw not in CRITICALITY_MAP:
            errs.append("درجه اهمیت نامعتبر است")

        status_raw = _norm(raw.get("status")).lower() or "active"
        if status_raw not in STATUS_MAP:
            errs.append("وضعیت نامعتبر است")

        if code:
            file_codes.setdefault(code, idx)

        row = ImportBatchRow(
            batch_id=batch.id, row_number=idx, raw=raw,
            is_valid=not errs, errors=errs,
        )
        db.add(row)
        if errs:
            errors += 1
        else:
            valid += 1
    return valid, errors


def _create_batch(db: Session, user: User, rows: list[dict], filename: str,
                  auto_create_lookups: bool, request: Request,
                  source: str = "excel") -> dict:
    """Create + validate an ImportBatch from parsed rows; returns preview payload."""
    batch = ImportBatch(
        filename=filename, status="pending", total_rows=len(rows),
        auto_create_lookups=auto_create_lookups, created_by=user.id,
    )
    db.add(batch)
    db.flush()

    existing_codes = {c for (c,) in db.query(Equipment.code).all()}
    valid, errors = _validate_batch(db, batch, rows, existing_codes)
    batch.valid_rows = valid
    batch.error_rows = errors
    batch.summary = {"valid": valid, "errors": errors, "source": source}

    audit.record(db, user_id=user.id, action="import.uploaded", entity_type="import_batch",
                 entity_id=batch.id,
                 new={"filename": batch.filename, "total": len(rows), "source": source},
                 request=request)
    db.commit()
    db.refresh(batch)

    return {
        "batch_id": batch.id,
        "status": "pending",
        "total_rows": batch.total_rows,
        "valid_rows": valid,
        "error_rows": errors,
        "rows": [
            {"row_number": r.row_number, "code": r.raw.get("code"),
             "name": r.raw.get("name"), "level": r.raw.get("level"),
             "is_valid": r.is_valid, "errors": r.errors or []}
            for r in batch.rows
        ],
    }


@router.post("")
async def upload_batch(
    request: Request,
    file: UploadFile = File(...),
    auto_create_lookups: bool = False,
    user: User = Depends(require("import.manage")),
    db: Session = Depends(get_db),
):
    fname = (file.filename or "").lower()
    if not fname.endswith((".xlsx", ".xls", ".md", ".markdown", ".txt")):
        raise HTTPException(
            status_code=400,
            detail="فقط فایل Excel (xlsx) یا Markdown/متنی (md / txt) مجاز است",
        )
    content = await file.read()
    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="حجم فایل بیش از ۱۰ مگابایت است")

    if fname.endswith((".md", ".markdown", ".txt")):
        try:
            text = content.decode("utf-8-sig")
        except UnicodeDecodeError:
            raise HTTPException(status_code=400, detail="فایل متنی باید با کدگذاری UTF-8 باشد")
        rows = parse_markdown_equipment(text)
        if not rows:
            raise HTTPException(
                status_code=400,
                detail="هیچ رکورد تجهیز در متن یافت نشد؛ جدول یا سرفصل‌ها قابل تشخیص نبودند",
            )
        return _create_batch(db, user, rows, file.filename or "upload.md",
                             auto_create_lookups, request, source="markdown")

    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="فایل Excel قابل خواندن نیست")

    ws = wb.active
    grid = [[c for c in row] for row in ws.iter_rows(values_only=True)]
    if not grid:
        raise HTTPException(status_code=400, detail="فایل خالی است")

    colmap = _map_header([_norm(c) for c in grid[0]])
    if "code" not in colmap.values() or "name" not in colmap.values():
        raise HTTPException(
            status_code=400,
            detail="ستون‌های «کد تجهیز» و «نام تجهیز» در سربرگ یافت نشدند؛ قالب نمونه را دانلود کنید",
        )

    rows: list[dict] = []
    for raw_row in grid[1:]:
        if raw_row is None or all(v is None or _norm(v) == "" for v in raw_row):
            continue
        rec = {field: _norm(raw_row[i]) if i < len(raw_row) else ""
               for i, field in colmap.items()}
        rows.append(rec)
    if not rows:
        raise HTTPException(status_code=400, detail="هیچ ردیف داده‌ای در فایل نیست")

    return _create_batch(db, user, rows, file.filename or "upload.xlsx",
                         auto_create_lookups, request, source="excel")


@router.post("/text")
def upload_text(
    body: TextImportIn,
    request: Request,
    user: User = Depends(require("import.manage")),
    db: Session = Depends(get_db),
):
    """Paste Markdown/plain-text equipment data → staging preview."""
    rows = parse_markdown_equipment(body.text)
    if not rows:
        raise HTTPException(
            status_code=400,
            detail="هیچ رکورد تجهیز در متن یافت نشد؛ جدول یا سرفصل‌ها قابل تشخیص نبودند",
        )
    return _create_batch(db, user, rows, body.filename or "markdown-paste.md",
                         body.auto_create_lookups, request, source="markdown")


@router.post("/{batch_id}/confirm")
def confirm_batch(batch_id: int, request: Request,
                  user: User = Depends(require("import.manage")),
                  db: Session = Depends(get_db)):
    batch = db.get(ImportBatch, batch_id)
    if not batch or batch.status != "pending":
        raise HTTPException(status_code=404, detail="بسته ورودی معتبر نیست")

    valid_rows = sorted(
        [r for r in batch.rows if r.is_valid],
        key=lambda r: {"equipment": 0, "subsystem": 1, "component": 2,
                       "subcomponent": 3}[LEVEL_MAP.get(_norm(r.raw.get("level")).lower(), "equipment")],
    )

    factories = {x.key: x.obj for x in []}  # built below
    f_map: dict[str, Factory] = {}
    for f in db.query(Factory).all():
        f_map[f.code.strip().lower()] = f
        f_map[f.name.strip().lower()] = f
    c_map: dict[str, EquipmentCategory] = {}
    for c in db.query(EquipmentCategory).all():
        c_map[c.code.strip().lower()] = c
        c_map[c.name.strip().lower()] = c

    code_to_equipment: dict[str, Equipment] = {}
    created = 0
    skipped: list[dict] = []

    for r in valid_rows:
        raw = r.raw
        level = LEVEL_MAP[_norm(raw.get("level")).lower()]
        code = _norm(raw.get("code"))
        if code in code_to_equipment or db.query(Equipment).filter(Equipment.code == code).one_or_none():
            skipped.append({"row": r.row_number, "code": code, "reason": "کد تکراری شد"})
            continue

        factory = category = None
        fk = _norm(raw.get("factory")).lower()
        ck = _norm(raw.get("category")).lower()
        if fk:
            factory = f_map.get(fk)
            if not factory and batch.auto_create_lookups:
                factory = Factory(code=f"IMP-{len(f_map)+1:03d}", name=_norm(raw.get("factory")),
                                  created_by=user.id)
                db.add(factory); db.flush()
                f_map[factory.name.strip().lower()] = factory
        if ck:
            category = c_map.get(ck)
            if not category and batch.auto_create_lookups:
                category = EquipmentCategory(code=f"IMPC-{len(c_map)+1:03d}",
                                             name=_norm(raw.get("category")),
                                             created_by=user.id)
                db.add(category); db.flush()
                c_map[category.name.strip().lower()] = category

        parent = None
        parent_code = _norm(raw.get("parent_code"))
        if parent_code:
            parent = code_to_equipment.get(parent_code) or (
                db.query(Equipment).filter(Equipment.code == parent_code).one_or_none()
            )
            if parent is None:
                skipped.append({"row": r.row_number, "code": code,
                                "reason": f"والد {parent_code} یافت نشد"})
                continue

        year_raw = _norm(raw.get("year"))
        year = int(float(year_raw)) if year_raw else None

        dyn = {}
        notes = _norm(raw.get("notes"))
        if notes:
            dyn["توضیحات ورود گروهی"] = notes

        e = Equipment(
            code=code, name=_norm(raw.get("name")), level=level,
            factory_id=factory.id if factory else (parent.factory_id if parent else None),
            category_id=category.id if category else (parent.category_id if parent else None),
            parent_id=parent.id if parent else None,
            location=_norm(raw.get("location")) or None,
            hall=_norm(raw.get("hall")) or None,
            dept=_norm(raw.get("dept")) or None,
            line=_norm(raw.get("line")) or None,
            position=_norm(raw.get("position")) or None,
            location_notes=_norm(raw.get("location_notes")) or None,
            component_type=_norm(raw.get("component_type")) or None,
            manufacturer=_norm(raw.get("manufacturer")) or None,
            model=_norm(raw.get("model")) or None,
            serial_number=_norm(raw.get("serial_number")) or None,
            year=year,
            criticality=CRITICALITY_MAP[_norm(raw.get("criticality")).lower() or "medium"],
            status=STATUS_MAP[_norm(raw.get("status")).lower() or "active"],
            dynamic_fields=dyn or None,
            created_by=user.id,
        )
        db.add(e)
        db.flush()
        code_to_equipment[code] = e
        r.created_equipment_id = e.id
        created += 1

    batch.status = "confirmed"
    batch.confirmed_at = utcnow()
    batch.summary = {"created": created, "skipped": skipped}

    audit.record(db, user_id=user.id, action="import.confirmed", entity_type="import_batch",
                 entity_id=batch.id, new={"created": created, "skipped": skipped},
                 request=request)
    db.commit()
    bus.publish("equipment.bulk_imported", {"batch_id": batch.id, "created": created})
    return {"ok": True, "created": created, "skipped": skipped}


@router.post("/{batch_id}/rollback")
def rollback_batch(batch_id: int, request: Request,
                   user: User = Depends(require("import.manage")),
                   db: Session = Depends(get_db)):
    batch = db.get(ImportBatch, batch_id)
    if not batch or batch.status != "confirmed":
        raise HTTPException(status_code=400, detail="فقط بسته‌های تأییدشده قابل بازگردانی هستند")

    removed = 0
    # Children first to respect FK ordering.
    rows = sorted(
        [r for r in batch.rows if r.created_equipment_id],
        key=lambda r: -({"equipment": 0, "subsystem": 1, "component": 2,
                         "subcomponent": 3}[LEVEL_MAP.get(_norm(r.raw.get("level")).lower(), "equipment")]),
    )
    from ..models import FileObject, MaintenancePlan

    # Capture ids, then release row references so FKs never block the delete.
    equipment_ids = [r.created_equipment_id for r in rows if r.created_equipment_id]
    for r in rows:
        r.created_equipment_id = None
    db.flush()

    for eid in equipment_ids:
        e = db.get(Equipment, eid)
        if e is None:
            continue
        db.query(MaintenancePlan).filter(
            MaintenancePlan.equipment_id == e.id
        ).delete(synchronize_session=False)
        db.query(MaintenancePlan).filter(
            MaintenancePlan.target_id == e.id
        ).update({"target_id": None}, synchronize_session=False)
        for f in db.query(FileObject).filter(
            FileObject.entity_type == "equipment", FileObject.entity_id == e.id
        ).all():
            db.delete(f)
        db.delete(e)
        removed += 1
    batch.status = "rolled_back"
    audit.record(db, user_id=user.id, action="import.rolled_back", entity_type="import_batch",
                 entity_id=batch.id, new={"removed": removed}, request=request)
    db.commit()
    bus.publish("equipment.bulk_rollback", {"batch_id": batch.id, "removed": removed})
    return {"ok": True, "removed": removed}


@router.get("/batches")
def list_batches(_: User = Depends(require("import.manage")), db: Session = Depends(get_db)):
    items = db.query(ImportBatch).order_by(ImportBatch.id.desc()).all()
    return {
        "items": [
            {"id": b.id, "filename": b.filename, "status": b.status,
             "total_rows": b.total_rows, "valid_rows": b.valid_rows,
             "error_rows": b.error_rows, "summary": b.summary,
             "created_at": b.created_at.isoformat() if b.created_at else None,
             "confirmed_at": b.confirmed_at.isoformat() if b.confirmed_at else None}
            for b in items
        ]
    }
