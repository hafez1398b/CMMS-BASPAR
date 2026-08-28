"""Inventory parts + external-warehouse Import Gateway (§23, §24).

The company warehouse is an external system; data enters through an Excel
gateway with the same discipline as equipment bulk-import:
Preview → Validation → Confirm → Rollback.
"""
from __future__ import annotations

import io

from fastapi import APIRouter, Depends, File, HTTPException, Request, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from pydantic import BaseModel, Field
from sqlalchemy.orm import Session

from .. import audit
from ..db import get_db, utcnow
from ..events import bus
from ..models import Equipment, ImportBatch, Part, Supplier, User
from ..rbac import require

router = APIRouter(prefix="/parts", tags=["parts"])

HEADER_ALIASES = {
    "code": "code", "کد قطعه": "code", "کد": "code",
    "name": "name", "نام قطعه": "name", "نام": "name",
    "unit": "unit", "واحد": "unit",
    "stock_qty": "stock_qty", "موجودی": "stock_qty",
    "min_qty": "min_qty", "حد سفارش": "min_qty", "حداقل موجودی": "min_qty",
    "criticality": "criticality", "درجه اهمیت": "criticality",
    "lead_time_days": "lead_time_days", "زمان تأمین": "lead_time_days",
    "supplier": "supplier", "تأمین‌کننده": "supplier",
    "alternative_part": "alternative_part", "قطعه جایگزین": "alternative_part",
    "equipment_code": "equipment_code", "کد تجهیز مرتبط": "equipment_code",
}

CRIT_MAP = {"low": "low", "کم": "low", "medium": "medium", "متوسط": "medium",
            "high": "high", "زیاد": "high", "بالا": "high",
            "critical": "critical", "بحرانی": "critical"}


class PartIn(BaseModel):
    code: str = Field(min_length=1, max_length=64)
    name: str = Field(min_length=2, max_length=190)
    unit: str | None = None
    stock_qty: float = 0
    min_qty: float = 0
    criticality: str = "medium"
    lead_time_days: int | None = None
    supplier: str | None = None
    supplier_id: int | None = None
    alternative_part: str | None = None
    equipment_id: int | None = None


def _resolve_supplier(db, name: str | None, supplier_id: int | None):
    """نام تأمین‌کننده را به رکورد موجود وصل می‌کند (بدون ساخت رکورد جدید —
    ایجاد تأمین‌کننده فقط از مسیر مدیریت تأمین‌کنندگان، برای پیشگیری از دیتای
    کثیف هنگام بارگذاری)."""
    if supplier_id and db.get(Supplier, supplier_id):
        return supplier_id
    if name:
        s = db.query(Supplier).filter(Supplier.name == name.strip()).one_or_none()
        if s:
            return s.id
    return None


def _out(p: Part) -> dict:
    return {
        "id": p.id, "code": p.code, "name": p.name, "unit": p.unit,
        "stock_qty": p.stock_qty, "min_qty": p.min_qty,
        "order_qty": p.order_qty,
        "criticality": p.criticality, "lead_time_days": p.lead_time_days,
        "supplier": p.supplier, "supplier_id": p.supplier_id,
        "supplier_name": p.supplier_ref.name if p.supplier_ref else None,
        "alternative_part": p.alternative_part,
        "equipment_id": p.equipment_id,
        "equipment_name": p.equipment.name if p.equipment else None,
        "low_stock": bool(p.stock_qty <= (p.min_qty or 0)),
    }


@router.get("")
def list_parts(q: str | None = None, low_stock: bool = False,
               _: User = Depends(require("parts.view")), db: Session = Depends(get_db)):
    query = db.query(Part)
    if q:
        like = f"%{q}%"
        query = query.filter(Part.code.ilike(like) | Part.name.ilike(like))
    items = query.order_by(Part.code).all()
    if low_stock:
        items = [p for p in items if p.stock_qty <= (p.min_qty or 0)]
    return {"items": [_out(p) for p in items]}


@router.post("", status_code=201)
def create_part(body: PartIn, request: Request,
                user: User = Depends(require("parts.manage")),
                db: Session = Depends(get_db)):
    if db.query(Part).filter(Part.code == body.code).one_or_none():
        raise HTTPException(status_code=409, detail="کد قطعه تکراری است")
    p = Part(**body.model_dump(), created_by=user.id)
    p.supplier_id = _resolve_supplier(db, body.supplier, body.supplier_id)
    db.add(p)
    db.flush()
    audit.record(db, user_id=user.id, action="part.created", entity_type="part",
                 entity_id=p.id, new=_out(p), request=request)
    db.commit()
    bus.publish("inventory.updated", {"part_id": p.id})
    return _out(p)


@router.put("/{pid}")
def update_part(pid: int, body: PartIn, request: Request,
                user: User = Depends(require("parts.manage")),
                db: Session = Depends(get_db)):
    p = db.get(Part, pid)
    if p is None:
        raise HTTPException(status_code=404, detail="قطعه یافت نشد")
    before = _out(p)
    data = body.model_dump()
    dup = db.query(Part).filter(Part.code == data["code"], Part.id != pid).one_or_none()
    if dup:
        raise HTTPException(status_code=409, detail="کد قطعه تکراری است")
    for k, v in data.items():
        setattr(p, k, v)
    p.supplier_id = _resolve_supplier(db, body.supplier, body.supplier_id)
    audit.record(db, user_id=user.id, action="part.updated", entity_type="part",
                 entity_id=p.id, old=before, new=_out(p), request=request)
    db.commit()
    bus.publish("inventory.updated", {"part_id": p.id})
    return _out(p)


@router.delete("/{pid}")
def delete_part(pid: int, request: Request,
                user: User = Depends(require("parts.manage")),
                db: Session = Depends(get_db)):
    p = db.get(Part, pid)
    if p is None:
        raise HTTPException(status_code=404, detail="قطعه یافت نشد")
    db.delete(p)
    audit.record(db, user_id=user.id, action="part.deleted", entity_type="part",
                 entity_id=pid, request=request)
    db.commit()
    return {"ok": True}


# ---------------------------------------------------------------------------
# Import Gateway (§23): Excel → preview/validate → confirm → rollback
# ---------------------------------------------------------------------------

TEMPLATE_COLUMNS = ["کد قطعه", "نام قطعه", "واحد", "موجودی", "حد سفارش",
                    "درجه اهمیت", "زمان تأمین", "تأمین‌کننده", "قطعه جایگزین",
                    "کد تجهیز مرتبط"]


@router.get("/import/template")
def parts_template(_: User = Depends(require("parts.view"))):
    wb = Workbook()
    ws = wb.active
    ws.title = "Parts"
    ws.append(TEMPLATE_COLUMNS)
    ws.append(["P-1001", "فیلتر روغن کمپرسور", "عدد", 4, 2, "زیاد", 21,
               "بازرگانی الف", "P-1002", "EQ-1001"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=parts-import-template.xlsx"})


def _norm(v):
    return "" if v is None else str(v).strip()


@router.post("/import")
async def import_parts(request: Request,
                       file: UploadFile = File(...),
                       user: User = Depends(require("parts.manage")),
                       db: Session = Depends(get_db)):
    if not (file.filename or "").lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="فقط فایل Excel مجاز است")
    content = await file.read()
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="فایل Excel قابل خواندن نیست")
    grid = [list(row) for row in wb.active.iter_rows(values_only=True)]
    if not grid:
        raise HTTPException(status_code=400, detail="فایل خالی است")

    colmap = {}
    for i, cell in enumerate(grid[0]):
        key = _norm(cell).lower()
        if key in HEADER_ALIASES:
            colmap[i] = HEADER_ALIASES[key]
    if "code" not in colmap.values() or "name" not in colmap.values():
        raise HTTPException(status_code=400,
                            detail="ستون‌های «کد قطعه» و «نام قطعه» الزامی هستند؛ قالب نمونه را دانلود کنید")

    existing = {c for (c,) in db.query(Part.code).all()}
    batch = ImportBatch(entity_type="parts", filename=file.filename, status="pending",
                        created_by=user.id)
    db.add(batch)
    db.flush()

    from ..models import ImportBatchRow
    seen, valid = set(), 0
    errors = 0
    eq_by_code = {e.code: e.id for e in
                  db.query(Equipment).filter(Equipment.deleted_at.is_(None)).all()}
    for idx, row in enumerate(grid[1:], start=2):
        if row is None or all(_norm(v) == "" for v in row):
            continue
        rec = {f: (_norm(row[i]) if i < len(row) else "") for i, f in colmap.items()}
        errs = []
        code = rec.get("code", "")
        if not code:
            errs.append("کد قطعه الزامی است")
        elif code in existing or code in seen:
            errs.append("کد قطعه تکراری است")
        if not rec.get("name"):
            errs.append("نام قطعه الزامی است")
        for numfield in ("stock_qty", "min_qty", "lead_time_days"):
            if rec.get(numfield):
                try:
                    float(rec[numfield])
                except ValueError:
                    errs.append(f"{numfield} عددی نیست")
        if rec.get("equipment_code") and rec["equipment_code"] not in eq_by_code:
            errs.append("تجهیز مرتبط یافت نشد")
        seen.add(code)
        db.add(ImportBatchRow(batch_id=batch.id, row_number=idx, raw=rec,
                              is_valid=not errs, errors=errs))
        valid += 0 if errs else 1
        errors += 1 if errs else 0

    batch.total_rows = valid + errors
    batch.valid_rows = valid
    batch.error_rows = errors
    audit.record(db, user_id=user.id, action="parts.import_uploaded",
                 entity_type="import_batch", entity_id=batch.id, request=request)
    db.commit()
    db.refresh(batch)
    return {
        "batch_id": batch.id, "total_rows": batch.total_rows,
        "valid_rows": valid, "error_rows": errors,
        "rows": [{"row_number": r.row_number, "code": r.raw.get("code"),
                  "name": r.raw.get("name"), "is_valid": r.is_valid,
                  "errors": r.errors or []} for r in batch.rows],
    }


@router.post("/import/{batch_id}/confirm")
def confirm_parts_import(batch_id: int, request: Request,
                         user: User = Depends(require("parts.manage")),
                         db: Session = Depends(get_db)):
    batch = db.get(ImportBatch, batch_id)
    if not batch or batch.entity_type != "parts" or batch.status != "pending":
        raise HTTPException(status_code=404, detail="بسته ورودی معتبر نیست")
    eq_by_code = {e.code: e.id for e in
                  db.query(Equipment).filter(Equipment.deleted_at.is_(None)).all()}
    created = 0
    for r in batch.rows:
        if not r.is_valid:
            continue
        raw = r.raw
        part = Part(
            code=raw.get("code", ""), name=raw.get("name", ""),
            unit=raw.get("unit") or None,
            stock_qty=float(raw.get("stock_qty") or 0),
            min_qty=float(raw.get("min_qty") or 0),
            criticality=CRIT_MAP.get(_norm(raw.get("criticality")).lower(), "medium"),
            lead_time_days=int(float(raw["lead_time_days"])) if raw.get("lead_time_days") else None,
            supplier=raw.get("supplier") or None,
            alternative_part=raw.get("alternative_part") or None,
            equipment_id=eq_by_code.get(raw.get("equipment_code")),
            import_batch_id=batch.id, created_by=user.id,
        )
        db.add(part)
        db.flush()
        r.created_equipment_id = None  # parts batch rows reuse the column? keep null
        created += 1
    batch.status = "confirmed"
    batch.confirmed_at = utcnow()
    batch.summary = {"created": created}
    audit.record(db, user_id=user.id, action="parts.import_confirmed",
                 entity_type="import_batch", entity_id=batch.id,
                 new={"created": created}, request=request)
    db.commit()
    bus.publish("inventory.updated", {"batch_id": batch.id, "created": created})
    return {"ok": True, "created": created}


@router.post("/import/{batch_id}/rollback")
def rollback_parts_import(batch_id: int, request: Request,
                          user: User = Depends(require("parts.manage")),
                          db: Session = Depends(get_db)):
    batch = db.get(ImportBatch, batch_id)
    if not batch or batch.entity_type != "parts" or batch.status != "confirmed":
        raise HTTPException(status_code=400, detail="فقط بسته‌های تأییدشده قابل بازگردانی هستند")
    removed = db.query(Part).filter(Part.import_batch_id == batch.id).delete()
    batch.status = "rolled_back"
    audit.record(db, user_id=user.id, action="parts.import_rolled_back",
                 entity_type="import_batch", entity_id=batch.id,
                 new={"removed": removed}, request=request)
    db.commit()
    return {"ok": True, "removed": removed}
