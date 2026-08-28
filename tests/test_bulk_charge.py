"""§6B MODULE EQUIPMENT — Bulk Data Charge Center: staging, SELEN mapping,
code decoding (§7), fuzzy duplicates, diff preview, commit, rollback."""
import io

import pytest
from openpyxl import Workbook


def _workbook(equipment_rows, specs=None, structure=None, parts=None,
              header=("کد تجهیز", "نام تجهیز", "کارخانه", "دسته", "سازنده",
                      "مدل", "سال ساخت", "درجه اهمیت", "وضعیت")):
    wb = Workbook()
    ws = wb.active
    ws.title = "تجهیزات"
    ws.append(list(header))
    for r in equipment_rows:
        ws.append(list(r))
    if specs:
        ws2 = wb.create_sheet("مشخصات فنی")
        ws2.append(["کد تجهیز", "نام مشخصه", "مقدار", "واحد"])
        for r in specs:
            ws2.append(list(r))
    if structure:
        ws3 = wb.create_sheet("ساختار")
        ws3.append(["کد تجهیز", "سطح", "کد والد", "نام", "سازنده", "مدل", "سریال"])
        for r in structure:
            ws3.append(list(r))
    if parts:
        ws4 = wb.create_sheet("قطعات")
        ws4.append(["کد تجهیز", "Part Number", "نام قطعه", "تعداد", "درجه اهمیت"])
        for r in parts:
            ws4.append(list(r))
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


@pytest.fixture(scope="module")
def mgr(client, admin_headers):
    from tests.test_phase1 import PW
    r = client.post("/api/users", headers=admin_headers, json={
        "username": "bcadmin", "full_name": "BC Admin", "password": PW,
        "role_names": ["technical_manager"],
    })
    assert r.status_code == 201
    t = client.post("/api/auth/login", json={"username": "bcadmin", "password": PW})
    return {"Authorization": f"Bearer {t.json()['access_token']}"}


def _upload(client, mgr, content, name="charge.xlsx"):
    return client.post("/api/equipment/bulk-charge/upload", headers=mgr,
                       files={"file": (name, content, "application/octet-stream")})


def test_upload_stages_data_without_touching_main_db(client, mgr):
    content = _workbook([
        ["B1BT-001", "مخزن بلندینگ ۱", "بسپار۱", "ترابری", "سازنده الف", "M1", 2018, "بحرانی", "فعال"],
        ["بازرسی خط تولید", "", "", "", "", "", "", "", ""],  # section row
        ["B1PT-002", "مخزن تولید ۲", "", "", "", "", "", "بالا", "فعال"],
    ])
    r = _upload(client, mgr, content)
    assert r.status_code == 201, r.text
    p = r.json()
    assert p["batch_id"] and p["total_rows"] >= 2
    # mapping suggestions present for every column
    assert len(p["mapping"]) == 9
    assert p["mapping"][0]["field"] == "code"

    # nothing committed yet
    eq = client.get("/api/equipment?level=all&q=B1BT-001", headers=mgr).json()
    assert eq["total"] == 0
    return p["batch_id"]


def test_code_decoding_and_preview_diff(client, mgr):
    content = _workbook([
        # code says بسپار۱ but factory column says بسپار۲ → must flag, not auto-fix
        ["B1BT-010", "مخزن ناهماهنگ", "بسپار۲", "ترابری", "", "", "", "متوسط", "فعال"],
        # factory column empty → infer بسپار۱ from code
        ["B1PT-011", "مخزن بدون ستون کارخانه", "", "ترابری", "", "", "", "کم", "فعال"],
        # unknown prefix → flag for admin
        ["XX-999", "پیشوند ناشناخته", "", "", "", "", "", "", "فعال"],
        # missing mandatory fields → rejected
        ["", "بدون کد", "بسپار۱", "ترابری", "", "", "", "", ""],
    ])
    r = _upload(client, mgr, content)
    bid = r.json()["batch_id"]
    pv = client.get(f"/api/equipment/bulk-charge/{bid}/preview", headers=mgr).json()
    assert pv["counts"]["rejected"] >= 1

    by_code = {row["code"]: row for row in pv["rows"] if row.get("code")}
    mismatch = by_code["B1BT-010"]
    assert mismatch["status"] == "conflict"
    assert any("ناهماهنگ" in e for e in mismatch["errors"])

    inferred = by_code["B1PT-011"]
    assert inferred["factory"] == "بسپار۱"  # SELEN filled from code prefix

    unknown = by_code["XX-999"]
    assert unknown["status"] == "conflict"
    assert any("پیشوند ناشناخته" in e for e in unknown["errors"])


def test_fuzzy_duplicate_flagged_not_auto_merged(client, mgr):
    # first, a committed record to be fuzzy-matched against
    r = client.post("/api/equipment", headers=mgr, json={
        "code": "FUZ-001", "name": "پمپ سانتریفیوژ گراندفوس CR-32", "level": "equipment",
        "factory_id": 1, "category_id": 1,
    })
    assert r.status_code == 201, r.text

    content = _workbook([
        ["", "پمپ سانتریفیوژ گراندفوس CR32", "", "", "", "", "", "", ""],  # no code, near-dup name
    ])
    # that row lacks mandatory code → rejected; craft one with a code instead
    content = _workbook([
        ["B1P-900", "پمپ سانتریفیوژ گراندفوس CR-32", "کارخانه مرکزی بسپار",
         "تأسیسات", "", "", "", "", ""],
    ])
    r = _upload(client, mgr, content)
    bid = r.json()["batch_id"]
    pv = client.get(f"/api/equipment/bulk-charge/{bid}/preview", headers=mgr).json()
    row = pv["rows"][0]
    assert row["status"] == "conflict"
    assert row["matched_equipment_id"] is not None
    assert any("احتمال تکراری" in e for e in row["errors"])


def test_update_vs_new_and_commit_rollback(client, admin_headers, mgr):
    # factory matching the B1 code prefix (§7) — base-data is admin-managed
    r = client.post("/api/factories", headers=admin_headers, json={
        "code": "FAC-B1", "name": "بسپار۱"})
    assert r.status_code in (201, 409)
    fid = next(f["id"] for f in client.get("/api/factories", headers=mgr).json()["items"]
               if f["code"] == "FAC-B1")
    # seed an existing equipment that an incoming row will UPDATE
    r = client.post("/api/equipment", headers=mgr, json={
        "code": "B1BT-EX1", "name": "تجهیز موجود", "level": "equipment",
        "factory_id": fid, "category_id": 1, "manufacturer": "قدیمی",
    })
    assert r.status_code == 201

    content = _workbook(
        [
            ["B1BT-EX1", "تجهیز موجود", "بسپار۱", "تأسیسات",
             "سازنده جدید", "MX", 2021, "زیاد", "فعال"],
            ["B1PT-NW1", "تجهیز کاملاً جدید", "بسپار۱", "تأسیسات",
             "", "", "", "متوسط", "فعال"],
        ],
        specs=[["B1PT-NW1", "توان", "7.5", "kW"], ["B1PT-NW1", "RPM", "1450", ""]],
        structure=[["B1PT-NW1", "زیرسیستم", "B1PT-NW1", "سیستم محرک", "", "", ""]],
        parts=[["B1PT-NW1", "PN-77", "یاتاقان ۶۲۰۴", 2, "زیاد"]],
    )
    r = _upload(client, mgr, content)
    bid = r.json()["batch_id"]
    pv = client.get(f"/api/equipment/bulk-charge/{bid}/preview", headers=mgr).json()
    assert pv["counts"]["update"] == 1 and pv["counts"]["new"] == 1
    assert pv["extra_sheets"]["specs"] == 2
    assert pv["extra_sheets"]["structure"] == 1
    assert pv["extra_sheets"]["parts"] == 1

    # commit applies specs/structure/parts too
    r = client.post(f"/api/equipment/bulk-charge/{bid}/commit", headers=mgr)
    assert r.status_code == 200, r.text
    s = r.json()
    assert s["created"] == 1 and s["updated"] == 1
    assert s["specs_applied"] == 2 and s["structure_created"] == 1 and s["parts_created"] == 1

    # main DB now reflects the charge (detail endpoint carries specs)
    eq = client.get("/api/equipment?level=all&q=B1PT-NW1", headers=mgr).json()["items"][0]
    eq = client.get(f"/api/equipment/{eq['id']}", headers=mgr).json()
    assert eq["technical_specs"]["توان (kW)"] == "7.5"
    parts = client.get(f"/api/equipment/{eq['id']}/parts", headers=mgr).json()
    assert any(p["part_number"] == "PN-77" for p in parts["items"])

    updated = client.get("/api/equipment?level=all&q=B1BT-EX1", headers=mgr).json()["items"][0]
    assert updated["manufacturer"] == "سازنده جدید"

    # rollback restores the update and removes the creation
    r = client.post(f"/api/equipment/bulk-charge/{bid}/rollback", headers=mgr)
    assert r.status_code == 200
    assert r.json()["removed"] == 1 and r.json()["restored"] == 1
    restored = client.get("/api/equipment?level=all&q=B1BT-EX1", headers=mgr).json()["items"][0]
    assert restored["manufacturer"] == "قدیمی"
    assert client.get("/api/equipment?level=all&q=B1PT-NW1", headers=mgr).json()["total"] == 0


def test_permissions_are_separate(client, admin_headers):
    """§6B: BulkCharge perms are distinct from Equipment.Create/Import."""
    from tests.test_phase1 import PW
    r = client.post("/api/users", headers=admin_headers, json={
        "username": "bctech", "full_name": "BC Tech", "password": PW,
        "role_names": ["technician"],
    })
    t = client.post("/api/auth/login", json={"username": "bctech", "password": PW})
    h = {"Authorization": f"Bearer {t.json()['access_token']}"}

    # technician holds bulk_charge.charge but NOT approve/rollback
    content = _workbook([["B1P-PRM1", "تست دسترسی", "بسپار۱",
                          "تأسیسات", "", "", "", "", "فعال"]])
    r = _upload(client, h, content)
    assert r.status_code == 201
    bid = r.json()["batch_id"]
    r = client.post(f"/api/equipment/bulk-charge/{bid}/commit", headers=h)
    assert r.status_code == 403


def test_rollback_guard_against_post_commit_edits(client, mgr):
    content = _workbook([
        ["B1BT-GRD1", "تجهیز گارد", "بسپار۱", "تأسیسات",
         "", "", "", "", "فعال"],
    ])
    bid = _upload(client, mgr, content).json()["batch_id"]
    client.post(f"/api/equipment/bulk-charge/{bid}/commit", headers=mgr)

    # someone edits the equipment after commit
    eq = client.get("/api/equipment?level=all&q=B1BT-GRD1", headers=mgr).json()["items"][0]
    r = client.patch(f"/api/equipment/{eq['id']}", headers=mgr, json={
        "version": eq["version"], "location": "تغییر بعد از commit"})
    assert r.status_code == 200

    r = client.post(f"/api/equipment/bulk-charge/{bid}/rollback", headers=mgr)
    assert r.status_code == 200
    assert r.json()["conflicts"] == 1 and r.json()["removed"] == 0


def test_pm_and_history_sheets_import(client, admin_headers, mgr):
    """Full legacy package: equipment + PM programs + repair history."""
    header = ("کد تجهیز", "نام تجهیز", "کارخانه", "دسته", "سازنده",
              "مدل", "سال ساخت", "درجه اهمیت", "وضعیت")
    wb = __import__("openpyxl").Workbook()
    ws = wb.active; ws.title = "تجهیزات"
    ws.append(list(header))
    ws.append(["B1PT-HX1", "مبدل حرارتی E-1", "بسپار۱", "تولیدی", "", "", "", "متوسط", "فعال"])
    ws5 = wb.create_sheet("برنامه نگهداری")
    ws5.append(["کد تجهیز", "عنوان فعالیت", "نوع فعالیت", "تناوب", "مجری", "مدت (دقیقه)", "آخرین اجرا"])
    ws5.append(["B1PT-HX1", "شستشوی لوله‌ها", "نظافت", "سه‌ماهه", "تیم نت", 90, "1404/02/10"])
    ws5.append(["B1PT-HX1", "بدون تناوب معتبر", "بازرسی", "هرگز", "", "", ""])
    ws6 = wb.create_sheet("سوابق تعمیرات")
    ws6.append(["کد تجهیز", "تاریخ", "نوع کار", "عنوان", "تکنسین", "مدت (دقیقه)", "هزینه (ریال)"])
    ws6.append(["B1PT-HX1", "1403/12/15", "تعمیر", "رفع گرفتگی لوله‌ها", "تکنسین ۱", 150, 5200000])
    ws6.append(["NO-SUCH-EQ", "1403/12/15", "تعمیر", "سابقه تجهیز ناموجود", "", "", ""])
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)

    r = client.post("/api/equipment/bulk-charge/upload", headers=mgr,
                    files={"file": ("legacy.xlsx", buf.getvalue(), "application/octet-stream")})
    assert r.status_code == 201
    bid = r.json()["batch_id"]
    assert r.json()["mapping"], "mapping suggestions returned"

    pv = client.get(f"/api/equipment/bulk-charge/{bid}/preview", headers=mgr).json()
    assert pv["extra_sheets"]["pm"] == 2 and pv["extra_sheets"]["history"] == 2
    # invalid interval + unknown equipment rows must be flagged
    flagged = [row for row in pv["rows"] if row["errors"]]
    assert any("تناوب" in "؛ ".join(row["errors"]) for row in flagged)
    assert any("یافت نشد" in "؛ ".join(row["errors"]) for row in flagged)

    cm = client.post(f"/api/equipment/bulk-charge/{bid}/commit", headers=mgr).json()
    assert cm["plans_created"] == 1 and cm["history_created"] == 1

    eq = client.get("/api/equipment?level=all&q=B1PT-HX1", headers=mgr).json()["items"][0]
    plans = client.get(f"/api/equipment/{eq['id']}/plans", headers=mgr).json()["items"]
    assert len(plans) == 1 and plans[0]["interval_code"] == "3monthly"
    assert plans[0]["next_due"] is not None  # computed from Jalali last execution
    hist = client.get(f"/api/equipment/{eq['id']}/history", headers=mgr).json()["items"]
    assert hist[0]["title"] == "رفع گرفتگی لوله‌ها" and hist[0]["work_type"] == "تعمیر"

    rb = client.post(f"/api/equipment/bulk-charge/{bid}/rollback", headers=mgr).json()
    assert rb["plans_removed"] == 1 and rb["history_removed"] == 1
    assert client.get(f"/api/equipment/{eq['id']}/plans", headers=mgr).json()["items"] == []


def test_flexible_date_parsing(client, mgr):
    from backend.app.modules.bulk_charge import _parse_flexible_date
    j = _parse_flexible_date("1404/05/26")
    g = _parse_flexible_date("2025-08-17")
    assert j is not None and g is not None and j.date() == g.date()
    assert _parse_flexible_date("not-a-date") is None
    assert _parse_flexible_date("") is None


def test_multisheet_one_equipment_per_sheet(client, admin_headers):
    """Real BASPAR format: several sheets, one equipment per sheet."""
    from openpyxl import Workbook
    wb = Workbook(); wb.remove(wb.active)
    rows = [("B1PF01", "دستگاه تزریق یک", "ماشین‌آلات تولید", "ماشین آلات", "سوله 1", "A"),
            ("B2AF01", "سانتریفیوژ فن", "ماشین‌آلات تولید", "*پیشگیری", "سوله 2", "C")]
    for code, name, cat, sect, hall, crit in rows:
        ws = wb.create_sheet(code)
        ws.append(["کد تجهیز", "دسته تجهیز", "خط تولید", "قسمت", "محل تجهیز",
                   "تاریخ نصب", "درجه اتوماسیون", "وضعیت تجهیز", "درجه اهمیت"])
        ws.append([code, name, cat, sect, hall, "1394/07/26", "اتوماتیک", "فعال", crit])
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    r = client.post("/api/equipment/bulk-charge/upload", headers=admin_headers,
                    files={"file": ("multi.xlsx", buf.getvalue(), "application/octet-stream")})
    assert r.status_code == 201
    bid = r.json()["batch_id"]
    assert r.json()["total_rows"] == 2
    pv = client.get(f"/api/equipment/bulk-charge/{bid}/preview", headers=admin_headers).json()
    assert pv["counts"]["new"] == 2
    cm = client.post(f"/api/equipment/bulk-charge/{bid}/commit", headers=admin_headers).json()
    assert cm["created"] == 2
    e = client.get("/api/equipment?level=equipment&q=B1PF01", headers=admin_headers).json()["items"][0]
    assert e["category"]["name"] == "ماشین‌آلات تولید"
    assert e["factory"]["name"] == "بسپار۱"
    assert e["criticality"] == "critical"  # A
    client.post(f"/api/equipment/bulk-charge/{bid}/rollback", headers=admin_headers)


def test_category_includes_production_machinery(client, admin_headers):
    from backend.app.db import SessionLocal
    from backend.app.models import EquipmentCategory
    with SessionLocal() as db:
        assert db.query(EquipmentCategory).filter(
            EquipmentCategory.name == "ماشین‌آلات تولید").count() >= 1


def test_history_creates_closed_workorders_with_assignment(client, mgr):
    """§4.3 + §5 سند نهایی: سوابق → دستورکار بسته‌شده با کد {تجهیز}-WO-n
    و تخصیص تکنسین طبق قانون نیروی انسانی."""
    # پرسنل قانون نیروی انسانی باید در دیتابیس تست وجود داشته باشند
    import importlib.util
    from pathlib import Path

    from backend.app.db import SessionLocal
    from backend.app.models import Role, User
    from backend.app.security import hash_password

    root = Path(__file__).resolve().parents[1]
    spec = importlib.util.spec_from_file_location(
        "seed_personnel", root / "scripts" / "seed_personnel.py")
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    with SessionLocal() as db:
        for username, full_name, roles, _n in mod.PERSONNEL:
            if db.query(User).filter(User.username == username).first():
                continue
            role_objs = db.query(Role).filter(Role.name.in_(roles)).all()
            db.add(User(username=username, full_name=full_name,
                        password_hash=hash_password("Baspar@1404"), roles=role_objs))
        db.commit()

    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active; ws.title = "تجهیزات"
    ws.append(["کد تجهیز", "نام تجهیز", "کارخانه", "دسته", "سازنده", "مدل",
               "سال ساخت", "درجه اهمیت", "وضعیت"])
    ws.append(["B1P-ASG1", "دستگاه تست انتساب", "بسپار۱", "ماشین‌آلات تولید",
               "", "", "", "زیاد", "فعال"])
    ws6 = wb.create_sheet("سوابق تعمیرات")
    ws6.append(["کد تجهیز", "تاریخ", "نوع کار", "عنوان", "اقدام تعمیراتی",
                "تکنسین", "مدت (دقیقه)", "هزینه (ریال)"])
    # برق → کاووسی (قبل از برج ۱۰)
    ws6.append(["B1P-ASG1", "1403/05/10", "EM", "خرابی سنسور پراکسی", "تعویض", "", "", ""])
    # مکانیک قبل از برج ۱۰ → بابایی
    ws6.append(["B1P-ASG1", "1403/06/20", "EM", "پک هیدرولیک", "تعویض", "", "", ""])
    # بعد از برج ۱۰ ۱۴۰۴ → معافی پور
    ws6.append(["B1P-ASG1", "1404/11/05", "PM", "پک هیدرولیک", "نظافت و بازرسی", "", "", ""])
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)

    r = client.post("/api/equipment/bulk-charge/upload", headers=mgr,
                    files={"file": ("asg.xlsx", buf.getvalue(), "application/octet-stream")})
    assert r.status_code == 201, r.text
    bid = r.json()["batch_id"]
    cm = client.post(f"/api/equipment/bulk-charge/{bid}/commit", headers=mgr).json()
    assert cm["history_created"] == 3, cm

    eq = client.get("/api/equipment?level=all&q=B1P-ASG1", headers=mgr).json()["items"][0]
    hist = client.get(f"/api/equipment/{eq['id']}/history", headers=mgr).json()["items"]
    assert len(hist) == 3
    techs = sorted((h.get("technician_name") or "") for h in hist)
    assert "احمد کاووسی" in techs and "نجات بابایی" in techs and "پوریا معافی پور" in techs

    # دستورکارهای بسته‌شده با کد استاندارد ایجاد شده‌اند
    wos = client.get(f"/api/work-orders?equipment_id={eq['id']}", headers=mgr).json()["items"]
    codes = sorted(w["code"] for w in wos)
    assert codes == ["B1P-ASG1-WO-01", "B1P-ASG1-WO-02", "B1P-ASG1-WO-03"]
    assert all(w["status"] == "closed" for w in wos)

    client.post(f"/api/equipment/bulk-charge/{bid}/rollback", headers=mgr)


def test_parts_sheet_with_inventory_fields(client, mgr):
    """§4.5 سند نهایی: کد قطعه/حد موجودی/مقدار سفارش/تأمین‌کننده."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active; ws.title = "تجهیزات"
    ws.append(["کد تجهیز", "نام تجهیز", "کارخانه", "دسته", "درجه اهمیت", "وضعیت"])
    ws.append(["B1P-PRT1", "تجهیز تست قطعات", "بسپار۱", "تولیدی", "متوسط", "فعال"])
    ws4 = wb.create_sheet("قطعات")
    ws4.append(["کد تجهیز", "کد قطعه", "نام قطعه", "موجودی فعلی", "حد موجودی",
                "مقدار سفارش", "تأمین‌کننده", "درجه اهمیت"])
    ws4.append(["B1P-PRT1", "SP-9001", "پک هیدرولیک", 2, 1, 4, "گستر اسپاد", "زیاد"])
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)

    r = client.post("/api/equipment/bulk-charge/upload", headers=mgr,
                    files={"file": ("prt.xlsx", buf.getvalue(), "application/octet-stream")})
    assert r.status_code == 201, r.text
    bid = r.json()["batch_id"]
    cm = client.post(f"/api/equipment/bulk-charge/{bid}/commit", headers=mgr).json()
    assert cm["parts_created"] == 1, cm

    parts = client.get("/api/parts?q=SP-9001", headers=mgr).json()["items"]
    p = next(x for x in parts if x["code"] == "SP-9001")
    assert p["min_qty"] == 1 and p["order_qty"] == 4 and p["supplier"] == "گستر اسپاد"

    client.post(f"/api/equipment/bulk-charge/{bid}/rollback", headers=mgr)
